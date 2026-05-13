from __future__ import annotations

import csv
import hashlib
import json
import math
import os
import re
import socket
import statistics
import subprocess
import sys
import time
import tracemalloc
import uuid
import shutil
from collections import Counter
from concurrent.futures import ThreadPoolExecutor
from dataclasses import dataclass
from itertools import product
from pathlib import Path
from typing import Dict, List, Sequence

try:
    import resource
except ModuleNotFoundError:  # pragma: no cover
    resource = None

try:
    from openpyxl import Workbook
except ModuleNotFoundError:  # pragma: no cover
    Workbook = None


@dataclass(frozen=True)
class MessageSample:
    payload: bytes
    label: int
    prediction: float | None = None


REPO_ROOT = Path(__file__).resolve().parents[3]
PREDICTIONS_PATH = REPO_ROOT / "SparkKafkaStreaming" / "predictions.csv"
NORMAL_RAW_PATH = REPO_ROOT / "Data" / "HRSS_normal_standard.csv"
ANOMALOUS_RAW_PATH = REPO_ROOT / "Data" / "HRSS_anomalous_standard.csv"
SPARK_RESOURCE_PROBE_PATH = Path(__file__).with_name("spark_resource_probe.py")
SPARK_RF_SCORE_PROBE_PATH = Path(__file__).with_name("spark_rf_score_probe.py")
CONSUMER_PATH = REPO_ROOT / "SparkKafkaStreaming" / "src" / "main" / "scala" / "spark" / "KafkaConsumerRF.scala"
ROUTING_PATH = REPO_ROOT / "SparkKafkaStreaming" / "src" / "main" / "scala" / "spark" / "KafkaAnomalyDetection.scala"
PDF_PATH = REPO_ROOT / "1-s2.0-S1568494624013875-main (2).pdf"
RESULTS_DIR = REPO_ROOT / "ModelandPerformanceAnalysis" / "results" / "bv_suite"
WORKBOOK_PATH = RESULTS_DIR / "BV_BigData_Test_Suite.xlsx"
JSON_PATH = RESULTS_DIR / "bv_bigdata_suite_results.json"
RF_PIPELINE_ARTIFACT_PATH = REPO_ROOT / "SparkKafkaStreaming" / "model_artifacts" / "random_forest_smote_standard_pipeline"
LIVE_METRICS_PATH = Path(
    os.getenv(
        "LIVE_METRICS_PATH",
        str(REPO_ROOT / "SparkKafkaStreaming" / "results" / "live_metrics" / "live_kafka_metrics.jsonl"),
    )
)
KAFKA_HOST = os.getenv("KAFKA_HOST", "localhost")
KAFKA_PORT = int(os.getenv("KAFKA_PORT", "9092"))
THRESHOLD_SWEEP_ANOMALY_VALUES = (0.8, 0.75, 0.7, 0.65, 0.6, 0.55)
THRESHOLD_SWEEP_NORMAL_VALUES = (0.2, 0.25, 0.3, 0.35, 0.4, 0.45)

PAPER_TESTS = [
    {
        "id": "PDF-4.4.1",
        "section": "4.4.1 / Table 9",
        "name": "Frequency of sensor data / data flow rate sweep",
        "variables": "data_flow_rate_ms = [10, 50, 100, 300, 400, 500]",
        "metrics": "tasks, time_s, avg_input_per_s, avg_process_per_s",
        "repo_before": "Missing. No runnable sweep in current repo.",
        "evidence": "KafkaConsumerRF.scala had a fixed batch interval and no rate sweep automation.",
    },
    {
        "id": "PDF-4.4.2",
        "section": "4.4.2 / Tables 10-11",
        "name": "Spark batch interval sweep",
        "variables": "batch_interval_s = [5, 10, 15, 20]",
        "metrics": "tasks, batch_count, jobs, storage_memory, peak_heap",
        "repo_before": "Partial only. Batch interval was fixed in streaming code.",
        "evidence": "No comparative batch-interval harness existed before the suite.",
    },
    {
        "id": "PDF-4.4.3",
        "section": "4.4.3 / Tables 12-13",
        "name": "Estimated delay times under different flow rates and batch intervals",
        "variables": "data_flow_rate_ms x batch_interval_s matrix",
        "metrics": "avg_input, avg_process, scheduling_delay_ms",
        "repo_before": "Missing. No end-to-end or scheduling-delay matrix automation.",
        "evidence": "Current repo had no runnable delay instrumentation around Kafka/Spark replay.",
    },
    {
        "id": "PDF-4.4.4",
        "section": "4.4.4 / Table 14",
        "name": "Spark parallelism / core-count sweep",
        "variables": "cores = [1, 2, 4, 16]",
        "metrics": "storage_memory, jvm_on_heap, jvm_off_heap, storage_on_heap",
        "repo_before": "Missing. local[*] existed but no core sweep test.",
        "evidence": "KafkaConsumerRF.scala and KafkaAnomalyDetection.scala were fixed local runs.",
    },
    {
        "id": "PDF-4.4.5",
        "section": "4.4.5 / Table 15",
        "name": "Kafka partition-count sweep",
        "variables": "partitions = [1, 3] with fixed core count",
        "metrics": "storage_memory, jvm_on_heap, tasks, task_time, elapsed_time, avg_input, avg_process",
        "repo_before": "Missing. No partition-count benchmark existed in the current repo.",
        "evidence": "No runnable topic partition benchmark or comparison script was present.",
    },
]


def env_flag(name: str, default: bool = False) -> bool:
    value = os.getenv(name)
    if value is None:
        return default
    return value.strip().lower() in {"1", "true", "yes", "on"}


def split_items(raw: str, separator: str = ",") -> List[str]:
    return [item.strip() for item in raw.split(separator) if item.strip()]


def env_list(name: str, default: str, separator: str = ",") -> List[str]:
    raw = os.getenv(name, default)
    return split_items(raw, separator=separator)


def env_int_list(name: str, default: str, separator: str = ",") -> List[int]:
    return [int(item) for item in env_list(name, default, separator=separator)]


def sanitize_label(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9]+", "-", value.strip())
    cleaned = cleaned.strip("-").lower()
    return cleaned or "profile"


def is_local_host(host: str) -> bool:
    normalized = host.strip().lower()
    return normalized in {"localhost", "127.0.0.1", "::1", "[::1]"}


def is_local_spark_master(master: str) -> bool:
    return master.strip().lower().startswith("local[")


def read_text(path: Path) -> str:
    return path.read_text(encoding="utf-8")


def extract_default(text: str, env_name: str, fallback: str = "unknown") -> str:
    match = re.search(rf'{re.escape(env_name)}",\s*"([^"]+)"', text)
    return match.group(1) if match else fallback


def extract_repo_evidence() -> Dict[str, str]:
    consumer_text = read_text(CONSUMER_PATH)
    routing_text = read_text(ROUTING_PATH)

    return {
        "spark_batch_interval_s": extract_default(consumer_text, "STREAM_BATCH_INTERVAL_SEC"),
        "spark_master_consumer": extract_default(consumer_text, "SPARK_MASTER", "local[*]"),
        "spark_master_routing": extract_default(routing_text, "SPARK_MASTER", "local[*]"),
        "input_topic": extract_default(consumer_text, "KAFKA_INPUT_TOPIC", "model-input"),
        "anomaly_threshold": extract_default(consumer_text, "ANOMALY_THRESHOLD", "0.8"),
        "normal_threshold": extract_default(consumer_text, "NORMAL_THRESHOLD", "0.2"),
        "anomaly_topic": extract_default(consumer_text, "ANOMALY_TOPIC", "anomalies3"),
        "normal_topic": extract_default(consumer_text, "NORMAL_TOPIC", "normal_data"),
        "uncertain_topic": extract_default(consumer_text, "UNCERTAIN_TOPIC", "uncertain_data"),
        "routing_policy": "route_uncertain" if "uncertainTopic" in consumer_text or "uncertainTopic" in routing_text else "unknown",
        "live_metrics_path": str(LIVE_METRICS_PATH),
        "pdf_reference": str(PDF_PATH.name),
    }


def load_prediction_samples(limit: int | None = None) -> List[MessageSample]:
    samples: List[MessageSample] = []
    with PREDICTIONS_PATH.open(newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        for idx, row in enumerate(reader):
            payload = json.dumps(row, separators=(",", ":"), sort_keys=True).encode("utf-8")
            samples.append(
                MessageSample(
                    payload=payload,
                    label=int(float(row["Labels"])),
                    prediction=float(row["predictions"]),
                )
            )
            if limit is not None and idx + 1 >= limit:
                break
    return samples


def load_rf_score_samples(limit: int | None = None) -> tuple[List[MessageSample] | None, Dict[str, object]]:
    spark_submit = resolve_spark_submit()
    if spark_submit is None:
        return None, {
            "status": "unavailable",
            "source_type": "rf_pipeline_replay",
            "reason": "SPARK_HOME/spark-submit not available for RF routing-score replay.",
        }
    if not SPARK_RF_SCORE_PROBE_PATH.exists():
        return None, {
            "status": "unavailable",
            "source_type": "rf_pipeline_replay",
            "reason": f"Missing RF routing-score probe script: {SPARK_RF_SCORE_PROBE_PATH}",
        }
    if not RF_PIPELINE_ARTIFACT_PATH.exists():
        return None, {
            "status": "unavailable",
            "source_type": "rf_pipeline_replay",
            "reason": f"Missing RF pipeline artifact: {RF_PIPELINE_ARTIFACT_PATH}",
        }

    output_path = RESULTS_DIR / f"rf_routing_scores_{uuid.uuid4().hex}.jsonl"
    stdout_log = output_path.with_suffix(".stdout.log")
    stderr_log = output_path.with_suffix(".stderr.log")
    env = os.environ.copy()
    env.update(
        {
            "ROUTING_SCORE_OUTPUT_JSONL": str(output_path),
            "RF_PIPELINE_ARTIFACT_PATH": str(RF_PIPELINE_ARTIFACT_PATH),
            "STREAM_SOURCE_NORMAL_CSV_PATH": env.get("STREAM_SOURCE_NORMAL_CSV_PATH", str(NORMAL_RAW_PATH)),
            "STREAM_SOURCE_ANOMALOUS_CSV_PATH": env.get("STREAM_SOURCE_ANOMALOUS_CSV_PATH", str(ANOMALOUS_RAW_PATH)),
        }
    )
    spark_args = [
        "--master",
        "local[*]",
        str(SPARK_RF_SCORE_PROBE_PATH),
    ]
    if spark_submit.suffix.lower() == ".cmd":
        invoke_script = output_path.with_suffix(".invoke.ps1")
        invoke_script.write_text(
            "\n".join(
                [
                    "$ErrorActionPreference = 'Stop'",
                    f"Set-Location {quote_powershell(str(REPO_ROOT))}",
                    "$argumentList = @(",
                    *[
                        f"    {quote_powershell(arg)}" + ("," if idx < len(spark_args) - 1 else "")
                        for idx, arg in enumerate(spark_args)
                    ],
                    ")",
                    f"$process = Start-Process -FilePath {quote_powershell(str(spark_submit))} -WorkingDirectory {quote_powershell(str(REPO_ROOT))} -ArgumentList $argumentList -RedirectStandardOutput {quote_powershell(str(stdout_log))} -RedirectStandardError {quote_powershell(str(stderr_log))} -PassThru -Wait -WindowStyle Hidden",
                    "exit $process.ExitCode",
                ]
            ),
            encoding="utf-8",
        )
        cmd = [
            "powershell.exe",
            "-NoProfile",
            "-ExecutionPolicy",
            "Bypass",
            "-File",
            str(invoke_script),
        ]
    else:
        cmd = [str(spark_submit), *spark_args]

    completed = subprocess.run(
        cmd,
        cwd=str(REPO_ROOT),
        env=env,
        capture_output=True,
        text=True,
        timeout=1200,
    )
    stdout_text = completed.stdout.strip()
    stderr_text = completed.stderr.strip()
    if stdout_log.exists():
        stdout_text = (stdout_text + "\n" + stdout_log.read_text(encoding="utf-8", errors="replace").strip()).strip()
    if stderr_log.exists():
        stderr_text = (stderr_text + "\n" + stderr_log.read_text(encoding="utf-8", errors="replace").strip()).strip()

    source_info: Dict[str, object] = {
        "source_type": "rf_pipeline_replay",
        "status": "failed",
        "model_artifact_path": str(RF_PIPELINE_ARTIFACT_PATH),
        "output_path": str(output_path),
        "command": subprocess.list2cmdline(cmd),
        "returncode": completed.returncode,
        "stdout_tail": truncate_text(stdout_text),
        "stderr_tail": truncate_text(stderr_text),
    }
    if completed.returncode != 0:
        source_info["reason"] = "RF routing-score probe returned a non-zero exit code."
        return None, source_info
    if not output_path.exists():
        source_info["reason"] = "RF routing-score probe completed without producing an output file."
        return None, source_info

    samples: List[MessageSample] = []
    with output_path.open(encoding="utf-8") as handle:
        for idx, line in enumerate(handle):
            row = json.loads(line)
            samples.append(
                MessageSample(
                    payload=b"",
                    label=int(row["label"]),
                    prediction=float(row["prediction"]),
                )
            )
            if limit is not None and idx + 1 >= limit:
                break

    source_info.update(
        {
            "status": "ok",
            "sample_count": len(samples),
            "reason": "RF routing scores were replay-scored from the current saved pipeline artifact.",
        }
    )
    return samples, source_info


def load_raw_source_rows(path: Path, required_label: int) -> List[Dict[str, str]]:
    rows: List[Dict[str, str]] = []
    with path.open(newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            label = int(float(row.get("Labels", "-1")))
            if label == required_label:
                rows.append(row)
    return rows


def interleave_rows(normal_rows: Sequence[Dict[str, str]], anomalous_rows: Sequence[Dict[str, str]]) -> List[Dict[str, str]]:
    merged: List[Dict[str, str]] = []
    normal_index = 0
    anomalous_index = 0
    normal_per_anomaly = max(1, len(normal_rows) // max(1, len(anomalous_rows)))

    while normal_index < len(normal_rows) or anomalous_index < len(anomalous_rows):
        emitted_normals = 0
        while normal_index < len(normal_rows) and emitted_normals < normal_per_anomaly:
            merged.append(normal_rows[normal_index])
            normal_index += 1
            emitted_normals += 1
        if anomalous_index < len(anomalous_rows):
            merged.append(anomalous_rows[anomalous_index])
            anomalous_index += 1
        if normal_index >= len(normal_rows):
            merged.extend(anomalous_rows[anomalous_index:])
            anomalous_index = len(anomalous_rows)
        if anomalous_index >= len(anomalous_rows):
            merged.extend(normal_rows[normal_index:])
            normal_index = len(normal_rows)

    return merged


def load_live_replay_samples(limit: int | None = None) -> List[MessageSample]:
    normal_rows = load_raw_source_rows(NORMAL_RAW_PATH, required_label=0)
    anomalous_rows = load_raw_source_rows(ANOMALOUS_RAW_PATH, required_label=1)
    ordered_rows = interleave_rows(normal_rows, anomalous_rows)
    samples: List[MessageSample] = []
    for idx, row in enumerate(ordered_rows):
        payload = json.dumps(row, separators=(",", ":"), sort_keys=True).encode("utf-8")
        samples.append(MessageSample(payload=payload, label=int(float(row["Labels"]))))
        if limit is not None and idx + 1 >= limit:
            break
    return samples


def route_band(sample: MessageSample, anomaly_threshold: float = 0.8, normal_threshold: float = 0.2) -> str:
    if sample.prediction is None:
        raise ValueError("Routing policy evaluation requires prediction scores.")
    if sample.prediction > anomaly_threshold:
        return "anomaly"
    if sample.prediction <= normal_threshold:
        return "normal"
    return "uncertain"


def evaluate_routing_policy(
    samples: Sequence[MessageSample],
    *,
    policy_name: str,
    anomaly_threshold: float = 0.8,
    normal_threshold: float = 0.2,
    route_uncertain: bool,
) -> Dict[str, float | str]:
    total = len(samples)
    positives = sum(1 for sample in samples if sample.label == 1)
    negatives = total - positives

    auto_tp = 0
    auto_fp = 0
    anomaly_topic_count = 0
    normal_topic_count = 0
    uncertain_band_count = 0
    review_topic_count = 0
    uncertain_positive = 0
    normal_positive = 0

    for sample in samples:
        band = route_band(sample, anomaly_threshold=anomaly_threshold, normal_threshold=normal_threshold)
        if band == "anomaly":
            anomaly_topic_count += 1
            if sample.label == 1:
                auto_tp += 1
            else:
                auto_fp += 1
        elif band == "normal":
            normal_topic_count += 1
            if sample.label == 1:
                normal_positive += 1
        else:
            uncertain_band_count += 1
            if route_uncertain:
                review_topic_count += 1
            if sample.label == 1:
                uncertain_positive += 1

    precision_alert = auto_tp / anomaly_topic_count if anomaly_topic_count else 0.0
    recall_alert = auto_tp / positives if positives else 0.0
    f1_alert = (
        2 * precision_alert * recall_alert / (precision_alert + recall_alert)
        if (precision_alert + recall_alert)
        else 0.0
    )
    fpr_alert = auto_fp / negatives if negatives else 0.0

    captured_with_review = auto_tp + (uncertain_positive if route_uncertain else 0)
    effective_recall_with_review = captured_with_review / positives if positives else 0.0
    residual_miss_rate_after_review = 1 - effective_recall_with_review if positives else 0.0

    return {
        "policy_name": policy_name,
        "anomaly_threshold": anomaly_threshold,
        "normal_threshold": normal_threshold,
        "anomaly_topic_count": anomaly_topic_count,
        "normal_topic_count": normal_topic_count,
        "uncertain_band_count": uncertain_band_count,
        "review_topic_count": review_topic_count,
        "uncertain_band_ratio": uncertain_band_count / total if total else 0.0,
        "review_load_ratio": review_topic_count / total if total else 0.0,
        "precision_on_auto_alert": precision_alert,
        "recall_on_auto_alert": recall_alert,
        "f1_on_auto_alert": f1_alert,
        "fpr_on_auto_alert": fpr_alert,
        "effective_recall_with_review": effective_recall_with_review,
        "residual_miss_rate_after_review": residual_miss_rate_after_review,
        "normal_bucket_positive_count": normal_positive,
    }


def build_policy_benchmark(samples: Sequence[MessageSample]) -> List[Dict[str, float | str]]:
    return [
        evaluate_routing_policy(
            samples,
            policy_name="current_drop_policy",
            route_uncertain=False,
        ),
        evaluate_routing_policy(
            samples,
            policy_name="route_uncertain_policy",
            route_uncertain=True,
        ),
    ]


def build_threshold_sweep(samples: Sequence[MessageSample]) -> tuple[List[Dict[str, object]], Dict[str, object]]:
    baseline = evaluate_routing_policy(
        samples,
        policy_name="baseline_route_uncertain_policy",
        anomaly_threshold=0.8,
        normal_threshold=0.2,
        route_uncertain=True,
    )
    rows: List[Dict[str, object]] = []
    for anomaly_threshold in THRESHOLD_SWEEP_ANOMALY_VALUES:
        for normal_threshold in THRESHOLD_SWEEP_NORMAL_VALUES:
            if normal_threshold >= anomaly_threshold:
                continue
            row = evaluate_routing_policy(
                samples,
                policy_name=f"route_uncertain_a{anomaly_threshold:.2f}_n{normal_threshold:.2f}",
                anomaly_threshold=anomaly_threshold,
                normal_threshold=normal_threshold,
                route_uncertain=True,
            )
            row.update(
                {
                    "delta_review_load_ratio_vs_baseline": row["review_load_ratio"] - baseline["review_load_ratio"],
                    "delta_effective_recall_with_review_vs_baseline": row["effective_recall_with_review"] - baseline["effective_recall_with_review"],
                    "delta_fpr_on_auto_alert_vs_baseline": row["fpr_on_auto_alert"] - baseline["fpr_on_auto_alert"],
                    "delta_precision_on_auto_alert_vs_baseline": row["precision_on_auto_alert"] - baseline["precision_on_auto_alert"],
                    "is_baseline": (
                        abs(float(row["anomaly_threshold"]) - float(baseline["anomaly_threshold"])) < 1e-9
                        and abs(float(row["normal_threshold"]) - float(baseline["normal_threshold"])) < 1e-9
                    ),
                }
            )
            rows.append(row)

    baseline_anomaly = float(baseline["anomaly_threshold"])
    baseline_normal = float(baseline["normal_threshold"])
    conservative_candidates = [
        row
        for row in rows
        if abs(float(row["anomaly_threshold"]) - baseline_anomaly) < 1e-9
        and float(row["effective_recall_with_review"]) >= float(baseline["effective_recall_with_review"]) - 0.10
        and float(row["fpr_on_auto_alert"]) <= float(baseline["fpr_on_auto_alert"]) + 0.01
        and float(row["precision_on_auto_alert"]) >= float(baseline["precision_on_auto_alert"]) - 0.01
    ]
    recommendation = min(
        conservative_candidates or rows,
        key=lambda row: (
            float(row["review_load_ratio"]),
            -float(row["effective_recall_with_review"]),
            float(row["fpr_on_auto_alert"]),
            -float(row["precision_on_auto_alert"]),
            abs(float(row["anomaly_threshold"]) - baseline_anomaly),
            abs(float(row["normal_threshold"]) - baseline_normal),
        ),
    )
    for row in rows:
        row["is_recommended"] = (
            abs(float(row["anomaly_threshold"]) - float(recommendation["anomaly_threshold"])) < 1e-9
            and abs(float(row["normal_threshold"]) - float(recommendation["normal_threshold"])) < 1e-9
        )

    rows.sort(key=lambda row: (float(row["normal_threshold"]), -float(row["anomaly_threshold"])))
    return rows, {
        "status": "available",
        "selection_policy": (
            "Keep the anomaly threshold at the deployed default when possible, then choose the lowest review load "
            "subject to <=0.10 absolute effective-recall loss and <=0.01 absolute FPR increase versus the baseline."
        ),
        "baseline_anomaly_threshold": baseline["anomaly_threshold"],
        "baseline_normal_threshold": baseline["normal_threshold"],
        "baseline_review_load_ratio": baseline["review_load_ratio"],
        "baseline_effective_recall_with_review": baseline["effective_recall_with_review"],
        "baseline_fpr_on_auto_alert": baseline["fpr_on_auto_alert"],
        "baseline_precision_on_auto_alert": baseline["precision_on_auto_alert"],
        "recommended_anomaly_threshold": recommendation["anomaly_threshold"],
        "recommended_normal_threshold": recommendation["normal_threshold"],
        "recommended_review_load_ratio": recommendation["review_load_ratio"],
        "recommended_effective_recall_with_review": recommendation["effective_recall_with_review"],
        "recommended_fpr_on_auto_alert": recommendation["fpr_on_auto_alert"],
        "recommended_precision_on_auto_alert": recommendation["precision_on_auto_alert"],
        "recommended_uncertain_band_ratio": recommendation["uncertain_band_ratio"],
        "review_load_reduction_ratio": baseline["review_load_ratio"] - recommendation["review_load_ratio"],
        "effective_recall_change": recommendation["effective_recall_with_review"] - baseline["effective_recall_with_review"],
        "fpr_change": recommendation["fpr_on_auto_alert"] - baseline["fpr_on_auto_alert"],
        "precision_change": recommendation["precision_on_auto_alert"] - baseline["precision_on_auto_alert"],
        "note": (
            "This is an operational threshold recommendation, not a model retraining result. "
            "It is intended to reduce uncertain/review load while keeping the RF alerting behavior close to the deployed baseline."
        ),
    }


def kafka_broker_reachable(host: str, port: int, timeout_s: float = 1.0) -> bool:
    try:
        with socket.create_connection((host, port), timeout=timeout_s):
            return True
    except OSError:
        return False


def parse_bootstrap_servers(bootstrap_servers: str) -> List[tuple[str, int]]:
    endpoints: List[tuple[str, int]] = []
    for item in split_items(bootstrap_servers, separator=","):
        if ":" in item:
            host, port_text = item.rsplit(":", 1)
            try:
                endpoints.append((host.strip(), int(port_text.strip())))
            except ValueError:
                continue
        else:
            endpoints.append((item.strip(), 9092))
    return endpoints


def summarize_bootstrap_reachability(bootstrap_servers: str) -> Dict[str, object]:
    endpoints = parse_bootstrap_servers(bootstrap_servers)
    attempts = []
    any_reachable = False
    for host, port in endpoints:
        reachable = kafka_broker_reachable(host, port)
        attempts.append(
            {
                "endpoint": f"{host}:{port}",
                "reachable": reachable,
            }
        )
        any_reachable = any_reachable or reachable
    return {
        "bootstrap_servers": bootstrap_servers,
        "reachable": any_reachable,
        "checked_endpoints": attempts,
    }


def bootstrap_servers_are_local(bootstrap_servers: str) -> bool:
    endpoints = parse_bootstrap_servers(bootstrap_servers)
    return bool(endpoints) and all(is_local_host(host) for host, _ in endpoints)


def classify_validation_scope(profiles: Sequence[Dict[str, object]]) -> str:
    if not profiles:
        return "unknown"
    all_local_spark = all(is_local_spark_master(str(profile.get("spark_master", ""))) for profile in profiles)
    all_local_kafka = all(bootstrap_servers_are_local(str(profile.get("kafka_bootstrap_servers", ""))) for profile in profiles)
    if all_local_spark and all_local_kafka:
        return "local_multi_profile_replay"
    if all_local_spark or all_local_kafka:
        return "hybrid_profile_matrix"
    return "distributed_cluster_replay"


def distributed_validation_note(mode: str, scope: str) -> str:
    if scope == "local_multi_profile_replay":
        if mode == "full_e2e_live_replay":
            return (
                "Each row reruns the full consumer-plus-producer replay on the same local Spark/Kafka stack "
                "under a different parameter profile. This is a local sensitivity matrix, not a multi-node deployment test."
            )
        return (
            "Each row reruns the direct Spark resource probe on the same local Spark/Kafka stack under a different "
            "parameter profile. This is a local sensitivity matrix, not a multi-node deployment test."
        )
    if scope == "hybrid_profile_matrix":
        return (
            "The profile set mixes local and non-local endpoints. Treat these rows as a configuration/reachability matrix "
            "unless every referenced Spark master and Kafka broker is actually available."
        )
    if mode == "full_e2e_live_replay":
        return (
            "Each row performs a full consumer-plus-producer live replay against the selected Spark master and Kafka "
            "bootstrap profile using unique topics, consumer groups, logs, and metrics files."
        )
    return (
        "Each row executes the direct Spark resource probe on the selected Spark master and records Kafka bootstrap "
        "reachability for the same deployment profile."
    )


def load_json_input(raw: str) -> object:
    candidate_path = Path(raw)
    if candidate_path.exists():
        return json.loads(candidate_path.read_text(encoding="utf-8"))
    return json.loads(raw)


def load_distributed_validation_profiles() -> List[Dict[str, object]]:
    local_only = env_flag("DISTRIBUTED_LOCAL_ONLY", False)
    explicit_matrix = os.getenv("DISTRIBUTED_VALIDATION_MATRIX_JSON", "").strip()
    if explicit_matrix and not local_only:
        payload = load_json_input(explicit_matrix)
        if isinstance(payload, dict):
            profiles = payload.get("profiles", [])
        elif isinstance(payload, list):
            profiles = payload
        else:
            raise ValueError("DISTRIBUTED_VALIDATION_MATRIX_JSON must resolve to a JSON list or an object with a 'profiles' array.")
        if not isinstance(profiles, list):
            raise ValueError("Distributed validation profiles must be a JSON array.")
        normalized: List[Dict[str, object]] = []
        for index, profile in enumerate(profiles, start=1):
            if not isinstance(profile, dict):
                raise ValueError(f"Distributed validation profile #{index} is not a JSON object.")
            spark_conf = profile.get("spark_conf") or {}
            if not isinstance(spark_conf, dict):
                raise ValueError(f"Distributed validation profile #{index} has a non-object 'spark_conf'.")
            spark_master = str(profile.get("spark_master", os.getenv("SPARK_MASTER", "local[*]"))).strip()
            kafka_bootstrap_servers = str(
                profile.get("kafka_bootstrap_servers", os.getenv("KAFKA_BOOTSTRAP_SERVERS", f"{KAFKA_HOST}:{KAFKA_PORT}"))
            ).strip()
            cores = int(profile.get("cores", 4))
            partitions = int(profile.get("partitions", 1))
            duration_s = int(profile.get("duration_s", 180))
            batch_interval_s = int(profile.get("batch_interval_s", 10))
            data_flow_rate_ms = int(profile.get("data_flow_rate_ms", 500))
            work_rounds = int(profile.get("work_rounds", 10))
            label = str(
                profile.get(
                    "label",
                    f"profile-{index}-{sanitize_label(spark_master)}-{sanitize_label(kafka_bootstrap_servers)}",
                )
            )
            normalized.append(
                {
                    "label": label,
                    "spark_master": spark_master,
                    "kafka_bootstrap_servers": kafka_bootstrap_servers,
                    "cores": cores,
                    "partitions": partitions,
                    "duration_s": duration_s,
                    "batch_interval_s": batch_interval_s,
                    "data_flow_rate_ms": data_flow_rate_ms,
                    "work_rounds": work_rounds,
                    "spark_conf": {str(key): str(value) for key, value in spark_conf.items()},
                }
            )
        return normalized

    if local_only:
        spark_masters = [os.getenv("LOCAL_VALIDATION_SPARK_MASTER", "local[*]").strip() or "local[*]"]
        kafka_profiles = [
            os.getenv("LOCAL_VALIDATION_BOOTSTRAP_SERVERS", f"{KAFKA_HOST}:{KAFKA_PORT}").strip()
            or f"{KAFKA_HOST}:{KAFKA_PORT}"
        ]
    else:
        spark_masters = env_list("DISTRIBUTED_SPARK_MASTERS", os.getenv("SPARK_MASTER", "local[*]"), separator=";")
        kafka_profiles = env_list(
            "DISTRIBUTED_KAFKA_BOOTSTRAP_LIST",
            os.getenv("KAFKA_BOOTSTRAP_SERVERS", f"{KAFKA_HOST}:{KAFKA_PORT}"),
            separator=";",
        )
    flow_rates = env_int_list("DISTRIBUTED_FLOW_RATES_MS", "500,1000")
    batch_intervals = env_int_list("DISTRIBUTED_BATCH_INTERVALS", "10")
    partitions_list = env_int_list("DISTRIBUTED_PARTITIONS", "1,3")
    cores_list = env_int_list("DISTRIBUTED_CORES", "4")
    duration_s = int(os.getenv("DISTRIBUTED_DURATION_S", "180"))
    work_rounds = int(os.getenv("DISTRIBUTED_WORK_ROUNDS", "10"))

    base_spark_conf: Dict[str, str] = {}
    for env_name, spark_key in (
        ("DISTRIBUTED_EXECUTOR_INSTANCES", "spark.executor.instances"),
        ("DISTRIBUTED_EXECUTOR_CORES", "spark.executor.cores"),
        ("DISTRIBUTED_EXECUTOR_MEMORY", "spark.executor.memory"),
        ("DISTRIBUTED_DRIVER_MEMORY", "spark.driver.memory"),
        ("DISTRIBUTED_MAX_CORES", "spark.cores.max"),
    ):
        value = os.getenv(env_name, "").strip()
        if value:
            base_spark_conf[spark_key] = value

    extra_conf_json = os.getenv("DISTRIBUTED_SPARK_CONF_JSON", "").strip()
    if extra_conf_json:
        extra_conf_payload = load_json_input(extra_conf_json)
        if not isinstance(extra_conf_payload, dict):
            raise ValueError("DISTRIBUTED_SPARK_CONF_JSON must resolve to a JSON object.")
        base_spark_conf.update({str(key): str(value) for key, value in extra_conf_payload.items()})

    profiles: List[Dict[str, object]] = []
    for index, (spark_master, kafka_bootstrap_servers, data_flow_rate_ms, batch_interval_s, partitions, cores) in enumerate(
        product(spark_masters, kafka_profiles, flow_rates, batch_intervals, partitions_list, cores_list),
        start=1,
    ):
        label = (
            f"profile-{index}-"
            f"{sanitize_label(spark_master)}-"
            f"{sanitize_label(kafka_bootstrap_servers)}-"
            f"f{data_flow_rate_ms}-b{batch_interval_s}-p{partitions}-c{cores}"
        )
        profiles.append(
            {
                "label": label,
                "spark_master": spark_master,
                "kafka_bootstrap_servers": kafka_bootstrap_servers,
                "cores": cores,
                "partitions": partitions,
                "duration_s": duration_s,
                "batch_interval_s": batch_interval_s,
                "data_flow_rate_ms": data_flow_rate_ms,
                "work_rounds": work_rounds,
                "spark_conf": dict(base_spark_conf),
            }
        )
    return profiles


def get_peak_rss_kib() -> float:
    if resource is not None:
        rss = float(resource.getrusage(resource.RUSAGE_SELF).ru_maxrss)
        if sys.platform == "darwin":
            return rss / 1024.0
        return rss

    if os.name == "nt":
        import ctypes
        from ctypes import wintypes

        class PROCESS_MEMORY_COUNTERS(ctypes.Structure):
            _fields_ = [
                ("cb", wintypes.DWORD),
                ("PageFaultCount", wintypes.DWORD),
                ("PeakWorkingSetSize", ctypes.c_size_t),
                ("WorkingSetSize", ctypes.c_size_t),
                ("QuotaPeakPagedPoolUsage", ctypes.c_size_t),
                ("QuotaPagedPoolUsage", ctypes.c_size_t),
                ("QuotaPeakNonPagedPoolUsage", ctypes.c_size_t),
                ("QuotaNonPagedPoolUsage", ctypes.c_size_t),
                ("PagefileUsage", ctypes.c_size_t),
                ("PeakPagefileUsage", ctypes.c_size_t),
            ]

        counters = PROCESS_MEMORY_COUNTERS()
        counters.cb = ctypes.sizeof(PROCESS_MEMORY_COUNTERS)
        process = ctypes.windll.kernel32.GetCurrentProcess()
        ok = ctypes.windll.psapi.GetProcessMemoryInfo(process, ctypes.byref(counters), counters.cb)
        if ok:
            return float(counters.PeakWorkingSetSize) / 1024.0

    return 0.0


def summarize_live_metrics() -> Dict[str, object]:
    summary = summarize_live_metrics_file(LIVE_METRICS_PATH, os.getenv("KAFKA_BOOTSTRAP_SERVERS", f"{KAFKA_HOST}:{KAFKA_PORT}"))
    summary.setdefault("broker_host", KAFKA_HOST)
    summary.setdefault("broker_port", KAFKA_PORT)
    if summary.get("status") == "missing":
        summary["detail"] = "Live metrics log not found. Start Kafka and the live consumer to populate it."
    elif summary.get("status") == "empty":
        summary["detail"] = "Live metrics file exists but contains no readable JSONL rows."
    return summary


def chunk_round_robin(items: Sequence[MessageSample], partitions: int) -> List[List[MessageSample]]:
    buckets = [[] for _ in range(partitions)]
    for idx, item in enumerate(items):
        buckets[idx % partitions].append(item)
    return [bucket for bucket in buckets if bucket]


def process_partition(batch: Sequence[MessageSample], work_rounds: int) -> Counter:
    counts = Counter()
    for sample in batch:
        digest = sample.payload
        for _ in range(work_rounds):
            digest = hashlib.sha256(digest).digest()
        counts["processed"] += 1
        counts["payload_bytes"] += len(digest)
    return counts


def run_stream_simulation(
    reservoir: Sequence[MessageSample],
    *,
    data_flow_rate_ms: int,
    batch_interval_s: int,
    duration_s: int,
    cores: int,
    partitions: int,
    work_rounds: int,
) -> Dict[str, float]:
    total_messages = max(1, int(duration_s * 1000 / data_flow_rate_ms))
    window_count = max(1, duration_s // batch_interval_s)
    batches: List[List[MessageSample]] = [[] for _ in range(window_count)]

    for index in range(total_messages):
        event_time_ms = index * data_flow_rate_ms
        if event_time_ms >= duration_s * 1000:
            break
        batch_index = int((event_time_ms / 1000) // batch_interval_s)
        if batch_index >= window_count:
            continue
        batches[batch_index].append(reservoir[index % len(reservoir)])

    tracemalloc.start()
    rss_before = get_peak_rss_kib()
    total_tasks = 0
    total_processed = 0
    nonempty_batches = 0
    batch_times_ms: List[float] = []
    scheduling_delays_ms: List[float] = []
    payload_bytes = 0
    actual_end_s = 0.0

    with ThreadPoolExecutor(max_workers=max(1, cores)) as executor:
        for batch_idx, batch in enumerate(batches):
            scheduled_start_s = batch_idx * batch_interval_s
            if not batch:
                scheduling_delays_ms.append(max(0.0, (actual_end_s - scheduled_start_s) * 1000.0))
                continue

            nonempty_batches += 1
            partitions_for_batch = chunk_round_robin(batch, max(1, partitions))
            total_tasks += len(partitions_for_batch)

            start = time.perf_counter()
            counters = list(executor.map(lambda part: process_partition(part, work_rounds), partitions_for_batch))
            processing_s = time.perf_counter() - start

            scheduled_delay_s = max(0.0, actual_end_s - scheduled_start_s)
            actual_start_s = scheduled_start_s + scheduled_delay_s
            actual_end_s = actual_start_s + processing_s

            batch_times_ms.append(processing_s * 1000.0)
            scheduling_delays_ms.append(scheduled_delay_s * 1000.0)
            for counter in counters:
                total_processed += counter["processed"]
                payload_bytes += counter["payload_bytes"]

    _, traced_peak = tracemalloc.get_traced_memory()
    tracemalloc.stop()
    rss_after = get_peak_rss_kib()
    rss_peak_kib = max(rss_before, rss_after)

    wall_elapsed_s = actual_end_s
    processing_elapsed_s = sum(batch_times_ms) / 1000.0 if batch_times_ms else 0.0
    avg_input_s = total_processed / duration_s if duration_s else 0.0
    avg_process_s = total_processed / processing_elapsed_s if processing_elapsed_s else 0.0
    mean_batch_ms = statistics.mean(batch_times_ms) if batch_times_ms else 0.0
    p95_batch_ms = quantile(batch_times_ms, 0.95)
    mean_delay_ms = statistics.mean(scheduling_delays_ms) if scheduling_delays_ms else 0.0
    max_delay_ms = max(scheduling_delays_ms) if scheduling_delays_ms else 0.0
    avg_payload_bytes = payload_bytes / total_processed if total_processed else 0.0

    return {
        "data_flow_rate_ms": data_flow_rate_ms,
        "batch_interval_s": batch_interval_s,
        "duration_s": duration_s,
        "cores": cores,
        "partitions": partitions,
        "messages": total_processed,
        "tasks": total_tasks,
        "batch_count": window_count,
        "jobs": nonempty_batches,
        "time_s": wall_elapsed_s,
        "avg_input_s": avg_input_s,
        "avg_process_s": avg_process_s,
        "mean_batch_ms": mean_batch_ms,
        "p95_batch_ms": p95_batch_ms,
        "mean_delay_ms": mean_delay_ms,
        "max_delay_ms": max_delay_ms,
        "peak_tracemalloc_kib": traced_peak / 1024.0,
        "peak_rss_mib": rss_peak_kib / 1024.0,
        "avg_payload_bytes": avg_payload_bytes,
    }


def quantile(values: Sequence[float], q: float) -> float:
    if not values:
        return 0.0
    ordered = sorted(values)
    if len(ordered) == 1:
        return ordered[0]
    position = (len(ordered) - 1) * q
    lower = math.floor(position)
    upper = math.ceil(position)
    if lower == upper:
        return ordered[lower]
    ratio = position - lower
    return ordered[lower] * (1 - ratio) + ordered[upper] * ratio


def resolve_spark_submit() -> Path | None:
    spark_home = os.getenv("SPARK_HOME")
    if spark_home:
        candidates = [Path(spark_home) / "bin" / "spark-submit.cmd", Path(spark_home) / "bin" / "spark-submit"]
        for candidate in candidates:
            if candidate.exists():
                return candidate
    return None


def resolve_runtime_jar() -> Path | None:
    preferred = os.getenv("EFFITRACK_RUNTIME_JAR", "").strip()
    if preferred:
        candidate = Path(preferred)
        if candidate.exists():
            return candidate

    target_dir = REPO_ROOT / "SparkKafkaStreaming" / "target" / "scala-2.11"
    named_candidates = [
        target_dir / "effitrack-live-thin.jar",
    ]
    named_candidates.extend(sorted(target_dir.glob("effitrack-live-thin-*.jar"), key=lambda path: path.stat().st_mtime, reverse=True))
    for candidate in named_candidates:
        if candidate.exists():
            return candidate
    return None


def resolve_kafka_topics_command() -> Path | None:
    kafka_home = os.getenv("KAFKA_HOME", "").strip()
    candidates: List[Path] = []
    if kafka_home:
        candidates.extend(
            [
                Path(kafka_home) / "bin" / "windows" / "kafka-topics.bat",
                Path(kafka_home) / "bin" / "kafka-topics.sh",
            ]
        )
    if os.name == "nt":
        candidates.extend(
            [
                Path(r"C:\kafka\bin\windows\kafka-topics.bat"),
                Path(r"C:\Kafka\bin\windows\kafka-topics.bat"),
            ]
        )
    else:
        for name in ("kafka-topics.sh", "kafka-topics"):
            resolved = shutil.which(name)
            if resolved:
                candidates.append(Path(resolved))

    for candidate in candidates:
        if candidate.exists():
            return candidate
    return None


def build_batch_command(executable: Path, args: Sequence[str]) -> List[str]:
    if executable.suffix.lower() == ".cmd":
        return ["cmd.exe", "/c", str(executable), *args]
    return [str(executable), *args]


def run_kafka_topics_command(
    kafka_topics_cmd: Path,
    args: Sequence[str],
    *,
    timeout: int = 120,
) -> subprocess.CompletedProcess[str]:
    cmd = build_batch_command(kafka_topics_cmd, list(args))
    return subprocess.run(
        cmd,
        cwd=str(REPO_ROOT),
        capture_output=True,
        text=True,
        timeout=timeout,
    )


def ensure_topic(
    kafka_topics_cmd: Path,
    *,
    bootstrap_servers: str,
    topic: str,
    partitions: int,
    replication_factor: int,
) -> Dict[str, object]:
    completed = run_kafka_topics_command(
        kafka_topics_cmd,
        [
            "--bootstrap-server",
            bootstrap_servers,
            "--create",
            "--if-not-exists",
            "--topic",
            topic,
            "--partitions",
            str(partitions),
            "--replication-factor",
            str(replication_factor),
        ],
    )
    return {
        "topic": topic,
        "partitions_requested": partitions,
        "returncode": completed.returncode,
        "stdout_tail": truncate_text(completed.stdout.strip()),
        "stderr_tail": truncate_text(completed.stderr.strip()),
    }


def describe_topic_partitions(kafka_topics_cmd: Path, *, bootstrap_servers: str, topic: str) -> int | None:
    completed = run_kafka_topics_command(
        kafka_topics_cmd,
        [
            "--bootstrap-server",
            bootstrap_servers,
            "--describe",
            "--topic",
            topic,
        ],
    )
    if completed.returncode != 0:
        return None
    count = 0
    for line in completed.stdout.splitlines():
        if "Partition:" in line:
            count += 1
    return count or None


def start_logged_process(
    executable: Path,
    args: Sequence[str],
    *,
    env: Dict[str, str],
    stdout_path: Path,
    stderr_path: Path,
) -> subprocess.Popen[str]:
    stdout_path.parent.mkdir(parents=True, exist_ok=True)
    stderr_path.parent.mkdir(parents=True, exist_ok=True)
    command = build_batch_command(executable, list(args))
    stdout_handle = stdout_path.open("w", encoding="utf-8", errors="replace")
    stderr_handle = stderr_path.open("w", encoding="utf-8", errors="replace")
    try:
        process = subprocess.Popen(
            command,
            cwd=str(REPO_ROOT),
            env=env,
            stdout=stdout_handle,
            stderr=stderr_handle,
            text=True,
        )
    finally:
        stdout_handle.close()
        stderr_handle.close()
    return process


def read_log_text(path: Path) -> str:
    if not path.exists():
        return ""
    return path.read_text(encoding="utf-8", errors="replace")


def wait_for_consumer_start(
    process: subprocess.Popen[str],
    *,
    stdout_path: Path,
    stderr_path: Path,
    timeout_s: int,
) -> tuple[bool, str, str]:
    deadline = time.time() + timeout_s
    while time.time() < deadline:
        stdout_text = read_log_text(stdout_path)
        stderr_text = read_log_text(stderr_path)
        if "KafkaConsumerRF started with inputTopic=" in stdout_text:
            return True, stdout_text, stderr_text
        if process.poll() is not None:
            return False, stdout_text, stderr_text
        if re.search(r"Exception|ERROR|error:", stdout_text + "\n" + stderr_text):
            return False, stdout_text, stderr_text
        time.sleep(1.0)
    return False, read_log_text(stdout_path), read_log_text(stderr_path)


def test_producer_completion(stdout_path: Path) -> bool:
    return "KafkaProducerRF replay completed." in read_log_text(stdout_path)


def stop_process_tree(process: subprocess.Popen[str] | None) -> None:
    if process is None:
        return
    if process.poll() is not None:
        return
    try:
        if os.name == "nt":
            subprocess.run(
                ["cmd.exe", "/c", "taskkill", "/PID", str(process.pid), "/T", "/F"],
                cwd=str(REPO_ROOT),
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
                timeout=30,
            )
        else:
            process.kill()
    except Exception:
        try:
            process.kill()
        except Exception:
            pass


def summarize_live_metrics_file(metrics_path: Path, bootstrap_servers: str) -> Dict[str, object]:
    reachability = summarize_bootstrap_reachability(bootstrap_servers)
    summary: Dict[str, object] = {
        "status": "missing",
        "metrics_file": str(metrics_path),
        "bootstrap_servers": bootstrap_servers,
        "broker_reachable": reachability["reachable"],
    }
    if not metrics_path.exists():
        summary["detail"] = "Metrics file not found."
        return summary

    rows: List[Dict[str, object]] = []
    with metrics_path.open(encoding="utf-8") as handle:
        for line in handle:
            line = line.strip()
            if not line:
                continue
            try:
                rows.append(json.loads(line))
            except json.JSONDecodeError:
                continue
    if not rows:
        summary["status"] = "empty"
        summary["detail"] = "Metrics file exists but contains no readable JSONL rows."
        return summary

    records_processed = sum(int(row.get("records_processed", 0)) for row in rows)
    weighted_latency_sum = sum(
        float(row.get("mean_e2e_latency_ms", 0.0)) * int(row.get("records_processed", 0))
        for row in rows
    )
    p95_batch_latencies = [float(row.get("p95_e2e_latency_ms", 0.0)) for row in rows]
    lag_values = [float(row.get("total_consumer_lag", 0.0)) for row in rows]
    batch_durations = [float(row.get("batch_duration_ms", 0.0)) for row in rows]
    anomaly_counts = sum(int(row.get("anomaly_count", 0)) for row in rows)
    normal_counts = sum(int(row.get("normal_count", 0)) for row in rows)
    uncertain_counts = sum(int(row.get("uncertain_count", 0)) for row in rows)

    summary.update(
        {
            "status": "available",
            "batches": len(rows),
            "records_processed": records_processed,
            "anomaly_count": anomaly_counts,
            "normal_count": normal_counts,
            "uncertain_count": uncertain_counts,
            "mean_e2e_latency_ms": weighted_latency_sum / records_processed if records_processed else 0.0,
            "p95_batch_e2e_latency_ms": quantile(p95_batch_latencies, 0.95),
            "mean_total_consumer_lag": statistics.mean(lag_values) if lag_values else 0.0,
            "max_total_consumer_lag": max(lag_values) if lag_values else 0.0,
            "mean_batch_duration_ms": statistics.mean(batch_durations) if batch_durations else 0.0,
            "policy_names": ", ".join(sorted({str(row.get("routing_policy", row.get("policy", "unknown"))) for row in rows})),
            "updated_at": time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(metrics_path.stat().st_mtime)),
        }
    )
    return summary


def direct_spark_metrics_enabled() -> bool:
    return SPARK_RESOURCE_PROBE_PATH.exists() and resolve_spark_submit() is not None


def truncate_text(value: str, limit: int = 1200) -> str:
    if len(value) <= limit:
        return value
    return value[:limit] + "\n...[truncated]..."


def quote_powershell(value: str) -> str:
    return "'" + value.replace("'", "''") + "'"


def run_spark_resource_benchmark(
    *,
    data_flow_rate_ms: int,
    batch_interval_s: int,
    duration_s: int,
    cores: int,
    partitions: int,
    work_rounds: int = 10,
    spark_master: str | None = None,
    kafka_bootstrap_servers: str | None = None,
    extra_spark_conf: Dict[str, str] | None = None,
    validation_label: str | None = None,
    diagnostics: List[Dict[str, object]] | None = None,
) -> Dict[str, float] | None:
    spark_submit = resolve_spark_submit()
    if spark_submit is None or not SPARK_RESOURCE_PROBE_PATH.exists():
        return None

    output_path = RESULTS_DIR / "spark_probe" / f"spark_resource_probe_{uuid.uuid4().hex}.json"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    stdout_log = output_path.with_suffix(".stdout.log")
    stderr_log = output_path.with_suffix(".stderr.log")
    env = os.environ.copy()
    env.update(
        {
            "BENCHMARK_DATA_FLOW_RATE_MS": str(data_flow_rate_ms),
            "BENCHMARK_BATCH_INTERVAL_S": str(batch_interval_s),
            "BENCHMARK_DURATION_S": str(duration_s),
            "BENCHMARK_CORES": str(cores),
            "BENCHMARK_PARTITIONS": str(partitions),
            "BENCHMARK_WORK_ROUNDS": str(work_rounds),
            "BENCHMARK_OUTPUT_JSON": str(output_path),
            "BENCHMARK_VALIDATION_LABEL": validation_label or "",
            "BENCHMARK_KAFKA_BOOTSTRAP_SERVERS": kafka_bootstrap_servers or os.getenv("KAFKA_BOOTSTRAP_SERVERS", ""),
            "STREAM_SOURCE_NORMAL_CSV_PATH": env.get("STREAM_SOURCE_NORMAL_CSV_PATH", str(NORMAL_RAW_PATH)),
            "STREAM_SOURCE_ANOMALOUS_CSV_PATH": env.get("STREAM_SOURCE_ANOMALOUS_CSV_PATH", str(ANOMALOUS_RAW_PATH)),
        }
    )

    selected_master = (spark_master or "").strip() or f"local[{cores}]"
    spark_args = [
        "--master",
        selected_master,
        str(SPARK_RESOURCE_PROBE_PATH),
    ]
    if extra_spark_conf:
        conf_args: List[str] = []
        for key, value in sorted(extra_spark_conf.items()):
            conf_args.extend(["--conf", f"{key}={value}"])
        spark_args = ["--master", selected_master, *conf_args, str(SPARK_RESOURCE_PROBE_PATH)]
    if spark_submit.suffix.lower() == ".cmd":
        invoke_script = output_path.with_suffix(".invoke.ps1")
        invoke_script.write_text(
            "\n".join(
                [
                    "$ErrorActionPreference = 'Stop'",
                    f"Set-Location {quote_powershell(str(REPO_ROOT))}",
                    "$argumentList = @(",
                    *[f"    {quote_powershell(arg)}" + ("," if idx < len(spark_args) - 1 else "") for idx, arg in enumerate(spark_args)],
                    ")",
                    f"$process = Start-Process -FilePath {quote_powershell(str(spark_submit))} -WorkingDirectory {quote_powershell(str(REPO_ROOT))} -ArgumentList $argumentList -RedirectStandardOutput {quote_powershell(str(stdout_log))} -RedirectStandardError {quote_powershell(str(stderr_log))} -PassThru -Wait -WindowStyle Hidden",
                    "exit $process.ExitCode",
                ]
            ),
            encoding="utf-8",
        )
        cmd = [
            "powershell.exe",
            "-NoProfile",
            "-ExecutionPolicy",
            "Bypass",
            "-File",
            str(invoke_script),
        ]
    else:
        cmd = [str(spark_submit), *spark_args]

    completed = subprocess.run(
        cmd,
        cwd=str(REPO_ROOT),
        env=env,
        capture_output=True,
        text=True,
        timeout=600,
    )
    stdout_text = completed.stdout.strip()
    stderr_text = completed.stderr.strip()
    if stdout_log.exists():
        stdout_text = (stdout_text + "\n" + stdout_log.read_text(encoding="utf-8", errors="replace").strip()).strip()
    if stderr_log.exists():
        stderr_text = (stderr_text + "\n" + stderr_log.read_text(encoding="utf-8", errors="replace").strip()).strip()

    diagnostic_row = {
        "validation_label": validation_label or "",
        "data_flow_rate_ms": data_flow_rate_ms,
        "batch_interval_s": batch_interval_s,
        "duration_s": duration_s,
        "cores": cores,
        "partitions": partitions,
        "spark_master": selected_master,
        "kafka_bootstrap_servers": kafka_bootstrap_servers or os.getenv("KAFKA_BOOTSTRAP_SERVERS", ""),
        "spark_conf": json.dumps(extra_spark_conf or {}, sort_keys=True),
        "command": subprocess.list2cmdline(cmd),
        "returncode": completed.returncode,
        "output_json_path": str(output_path),
        "output_json_exists": output_path.exists(),
        "stdout_tail": truncate_text(stdout_text),
        "stderr_tail": truncate_text(stderr_text),
    }
    if completed.returncode != 0:
        if diagnostics is not None:
            diagnostic_row["status"] = "failed_returncode"
            diagnostics.append(diagnostic_row)
        return None
    if not output_path.exists():
        if diagnostics is not None:
            diagnostic_row["status"] = "missing_output_json"
            diagnostics.append(diagnostic_row)
        return None
    try:
        result = json.loads(output_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        if diagnostics is not None:
            diagnostic_row["status"] = "invalid_json_output"
            diagnostic_row["json_error"] = str(exc)
            diagnostics.append(diagnostic_row)
        return None

    result["spark_master"] = selected_master
    result["kafka_bootstrap_servers"] = kafka_bootstrap_servers or os.getenv("KAFKA_BOOTSTRAP_SERVERS", "")
    result["validation_label"] = validation_label or ""
    result["spark_conf"] = dict(extra_spark_conf or {})
    if diagnostics is not None:
        diagnostic_row["status"] = "ok"
        diagnostics.append(diagnostic_row)
    return result


def overlay_direct_metrics(base_row: Dict[str, float], direct_row: Dict[str, float] | None) -> Dict[str, float]:
    if not direct_row:
        return base_row

    merged = dict(base_row)
    merged.update(
        {
            "messages": int(direct_row["messages"]),
            "tasks": int(direct_row["tasks"]),
            "batch_count": int(direct_row["batch_count"]),
            "jobs": int(direct_row["jobs"]),
            "time_s": float(direct_row["elapsed_time_s"]),
            "task_time_s": float(direct_row["task_time_s"]),
            "avg_input_s": float(direct_row["avg_input_s"]),
            "avg_process_s": float(direct_row["avg_process_s"]),
            "mean_batch_ms": float(direct_row["mean_batch_ms"]),
            "mean_delay_ms": float(direct_row["mean_delay_ms"]),
            "storage_memory_kib": float(direct_row["storage_memory_kib"]),
            "peak_jvm_on_heap_mib": float(direct_row["peak_jvm_on_heap_mib"]),
            "peak_jvm_off_heap_mib": float(direct_row["peak_jvm_off_heap_mib"]),
            "peak_storage_on_heap_mib": float(direct_row["peak_storage_on_heap_mib"]),
        }
    )
    return merged


def build_distributed_probe_validation_matrix() -> tuple[Dict[str, object], List[Dict[str, object]], List[Dict[str, object]]]:
    if not env_flag("ENABLE_DISTRIBUTED_VALIDATION", False):
        return (
            {
                "status": "disabled",
                "mode": "spark_probe_plus_broker_reachability",
                "reason": "ENABLE_DISTRIBUTED_VALIDATION was not set.",
            },
            [],
            [],
        )

    if not direct_spark_metrics_enabled():
        return (
            {
                "status": "unavailable",
                "mode": "spark_probe_plus_broker_reachability",
                "reason": "spark-submit/SPARK_RESOURCE_PROBE_PATH is unavailable, so distributed validation cannot run.",
            },
            [],
            [],
        )

    profiles = load_distributed_validation_profiles()
    if not profiles:
        return (
            {
                "status": "empty",
                "mode": "spark_probe_plus_broker_reachability",
                "reason": "No distributed validation profiles were resolved from the environment.",
            },
            [],
            [],
        )

    diagnostics: List[Dict[str, object]] = []
    rows: List[Dict[str, object]] = []
    for profile in profiles:
        reachability = summarize_bootstrap_reachability(str(profile["kafka_bootstrap_servers"]))
        direct_row = run_spark_resource_benchmark(
            data_flow_rate_ms=int(profile["data_flow_rate_ms"]),
            batch_interval_s=int(profile["batch_interval_s"]),
            duration_s=int(profile["duration_s"]),
            cores=int(profile["cores"]),
            partitions=int(profile["partitions"]),
            work_rounds=int(profile["work_rounds"]),
            spark_master=str(profile["spark_master"]),
            kafka_bootstrap_servers=str(profile["kafka_bootstrap_servers"]),
            extra_spark_conf=dict(profile["spark_conf"]),
            validation_label=str(profile["label"]),
            diagnostics=diagnostics,
        )

        row: Dict[str, object] = {
            "validation_label": profile["label"],
            "spark_master": profile["spark_master"],
            "kafka_bootstrap_servers": profile["kafka_bootstrap_servers"],
            "broker_reachable": reachability["reachable"],
            "checked_endpoint_count": len(reachability["checked_endpoints"]),
            "data_flow_rate_ms": profile["data_flow_rate_ms"],
            "batch_interval_s": profile["batch_interval_s"],
            "duration_s": profile["duration_s"],
            "cores": profile["cores"],
            "partitions": profile["partitions"],
            "spark_conf": json.dumps(profile["spark_conf"], sort_keys=True),
        }

        if direct_row is None:
            diagnostic = diagnostics[-1] if diagnostics else {}
            row.update(
                {
                    "status": diagnostic.get("status", "failed"),
                    "returncode": diagnostic.get("returncode", ""),
                    "output_json_exists": diagnostic.get("output_json_exists", False),
                    "time_s": "",
                    "mean_batch_ms": "",
                    "mean_delay_ms": "",
                    "peak_jvm_on_heap_mib": "",
                    "peak_jvm_off_heap_mib": "",
                }
            )
        else:
            row.update(
                {
                    "status": "ok",
                    "returncode": 0,
                    "output_json_exists": True,
                    "time_s": float(direct_row["elapsed_time_s"]),
                    "tasks": int(direct_row["tasks"]),
                    "jobs": int(direct_row["jobs"]),
                    "avg_input_s": float(direct_row["avg_input_s"]),
                    "avg_process_s": float(direct_row["avg_process_s"]),
                    "mean_batch_ms": float(direct_row["mean_batch_ms"]),
                    "mean_delay_ms": float(direct_row["mean_delay_ms"]),
                    "peak_jvm_on_heap_mib": float(direct_row["peak_jvm_on_heap_mib"]),
                    "peak_jvm_off_heap_mib": float(direct_row["peak_jvm_off_heap_mib"]),
                    "storage_memory_kib": float(direct_row["storage_memory_kib"]),
                }
            )
        rows.append(row)

    success_count = sum(1 for row in rows if row.get("status") == "ok")
    broker_ready_count = sum(1 for row in rows if row.get("broker_reachable"))
    validation_scope = classify_validation_scope(profiles)
    summary = {
        "status": "available",
        "mode": "spark_probe_plus_broker_reachability",
        "validation_scope": validation_scope,
        "profile_count": len(rows),
        "successful_probe_count": success_count,
        "failed_probe_count": len(rows) - success_count,
        "broker_reachable_profile_count": broker_ready_count,
        "spark_master_count": len({str(row["spark_master"]) for row in rows}),
        "kafka_profile_count": len({str(row["kafka_bootstrap_servers"]) for row in rows}),
        "note": distributed_validation_note("spark_probe_plus_broker_reachability", validation_scope),
    }
    return summary, rows, diagnostics


def spark_conf_args(conf: Dict[str, str]) -> List[str]:
    args: List[str] = []
    for key, value in sorted(conf.items()):
        args.extend(["--conf", f"{key}={value}"])
    return args


def run_distributed_live_replay_profile(
    profile: Dict[str, object],
    *,
    runtime_jar: Path,
    spark_submit: Path,
    spark_packages: str,
    kafka_topics_cmd: Path | None,
    run_token: str,
) -> tuple[Dict[str, object], Dict[str, object]]:
    label = str(profile["label"])
    bootstrap_servers = str(profile["kafka_bootstrap_servers"])
    spark_master = str(profile["spark_master"])
    producer_spark_master = str(profile.get("producer_spark_master", "local[1]"))
    cores = int(profile["cores"])
    partitions = int(profile["partitions"])
    duration_s = int(profile["duration_s"])
    batch_interval_s = int(profile["batch_interval_s"])
    data_flow_rate_ms = int(profile["data_flow_rate_ms"])
    work_rounds = int(profile["work_rounds"])
    extra_spark_conf = dict(profile.get("spark_conf", {}))
    reachability = summarize_bootstrap_reachability(bootstrap_servers)

    safe_label = sanitize_label(label)
    topic_suffix = f"{safe_label}-{run_token}"
    input_topic = f"effitrack-dv-{topic_suffix}-input"
    anomaly_topic = f"effitrack-dv-{topic_suffix}-anomaly"
    normal_topic = f"effitrack-dv-{topic_suffix}-normal"
    uncertain_topic = f"effitrack-dv-{topic_suffix}-uncertain"
    consumer_group = f"effitrack-dv-{topic_suffix}-group"
    metrics_dir = RESULTS_DIR / "distributed_validation"
    consumer_stdout = metrics_dir / f"{topic_suffix}_consumer_stdout.log"
    consumer_stderr = metrics_dir / f"{topic_suffix}_consumer_stderr.log"
    producer_stdout = metrics_dir / f"{topic_suffix}_producer_stdout.log"
    producer_stderr = metrics_dir / f"{topic_suffix}_producer_stderr.log"
    metrics_path = metrics_dir / f"{topic_suffix}_live_metrics.jsonl"
    metrics_dir.mkdir(parents=True, exist_ok=True)
    metrics_path.touch(exist_ok=True)

    requested_messages = max(1, int(duration_s * 1000 / max(1, data_flow_rate_ms)))
    topic_admin_required = not env_flag("DISTRIBUTED_SKIP_TOPIC_CREATE", False)
    replication_factor = int(os.getenv("DISTRIBUTED_KAFKA_REPLICATION_FACTOR", "1"))
    consumer_start_timeout_s = int(os.getenv("DISTRIBUTED_CONSUMER_START_TIMEOUT_SEC", os.getenv("CONSUMER_START_TIMEOUT_SEC", "180")))
    consumer_warmup_s = int(os.getenv("DISTRIBUTED_CONSUMER_WARMUP_SEC", "5"))
    post_run_wait_s = int(os.getenv("DISTRIBUTED_POST_RUN_WAIT_SEC", os.getenv("POST_RUN_WAIT_SEC", "8")))
    producer_timeout_s = int(os.getenv("DISTRIBUTED_PRODUCER_TIMEOUT_SEC", str(max(600, duration_s * 3))))

    row: Dict[str, object] = {
        "validation_label": label,
        "mode": "full_e2e_live_replay",
        "spark_master": spark_master,
        "producer_spark_master": producer_spark_master,
        "kafka_bootstrap_servers": bootstrap_servers,
        "broker_reachable": reachability["reachable"],
        "checked_endpoint_count": len(reachability["checked_endpoints"]),
        "data_flow_rate_ms": data_flow_rate_ms,
        "batch_interval_s": batch_interval_s,
        "duration_s": duration_s,
        "cores": cores,
        "partitions": partitions,
        "work_rounds": work_rounds,
        "spark_conf": json.dumps(extra_spark_conf, sort_keys=True),
        "input_topic": input_topic,
        "anomaly_topic": anomaly_topic,
        "normal_topic": normal_topic,
        "uncertain_topic": uncertain_topic,
        "consumer_group": consumer_group,
        "requested_messages": requested_messages,
        "metrics_file": str(metrics_path),
    }

    diagnostic: Dict[str, object] = {
        "validation_label": label,
        "mode": "full_e2e_live_replay",
        "spark_master": spark_master,
        "producer_spark_master": producer_spark_master,
        "kafka_bootstrap_servers": bootstrap_servers,
        "input_topic": input_topic,
        "metrics_file": str(metrics_path),
        "consumer_stdout_log": str(consumer_stdout),
        "consumer_stderr_log": str(consumer_stderr),
        "producer_stdout_log": str(producer_stdout),
        "producer_stderr_log": str(producer_stderr),
        "spark_conf": json.dumps(extra_spark_conf, sort_keys=True),
    }

    if not reachability["reachable"]:
        row["status"] = "broker_unreachable"
        diagnostic["status"] = "broker_unreachable"
        return row, diagnostic

    if topic_admin_required and kafka_topics_cmd is None:
        row["status"] = "topic_admin_unavailable"
        diagnostic["status"] = "topic_admin_unavailable"
        diagnostic["reason"] = "Kafka topic admin command was not found."
        return row, diagnostic

    topic_create_results: List[Dict[str, object]] = []
    actual_input_partitions: int | None = None
    if kafka_topics_cmd is not None:
        for topic_name in (input_topic, anomaly_topic, normal_topic, uncertain_topic):
            topic_create_results.append(
                ensure_topic(
                    kafka_topics_cmd,
                    bootstrap_servers=bootstrap_servers,
                    topic=topic_name,
                    partitions=partitions,
                    replication_factor=replication_factor,
                )
            )
        actual_input_partitions = describe_topic_partitions(
            kafka_topics_cmd,
            bootstrap_servers=bootstrap_servers,
            topic=input_topic,
        )

    row["actual_input_topic_partitions"] = actual_input_partitions if actual_input_partitions is not None else ""
    diagnostic["topic_create_results"] = json.dumps(topic_create_results, ensure_ascii=False)

    env = os.environ.copy()
    env.update(
        {
            "KAFKA_BOOTSTRAP_SERVERS": bootstrap_servers,
            "KAFKA_INPUT_TOPIC": input_topic,
            "ANOMALY_TOPIC": anomaly_topic,
            "NORMAL_TOPIC": normal_topic,
            "UNCERTAIN_TOPIC": uncertain_topic,
            "KAFKA_CONSUMER_GROUP": consumer_group,
            "KAFKA_AUTO_OFFSET_RESET": os.getenv("DISTRIBUTED_KAFKA_AUTO_OFFSET_RESET", "earliest"),
            "LIVE_METRICS_PATH": str(metrics_path),
            "SPARK_MASTER": spark_master,
            "STREAM_BATCH_INTERVAL_SEC": str(batch_interval_s),
            "REPLAY_FLOW_RATE_MS": str(data_flow_rate_ms),
            "MAX_MESSAGES": str(requested_messages),
            "BENCHMARK_WORK_ROUNDS": str(work_rounds),
        }
    )

    consumer_args = [
        "--master",
        spark_master,
        *spark_conf_args(extra_spark_conf),
        "--packages",
        spark_packages,
        "--class",
        "KafkaConsumerRF",
        str(runtime_jar),
    ]
    producer_args = [
        "--master",
        producer_spark_master,
        "--packages",
        spark_packages,
        "--class",
        "KafkaProducerRF",
        str(runtime_jar),
    ]

    consumer_process: subprocess.Popen[str] | None = None
    producer_process: subprocess.Popen[str] | None = None
    try:
        consumer_process = start_logged_process(
            spark_submit,
            consumer_args,
            env=env,
            stdout_path=consumer_stdout,
            stderr_path=consumer_stderr,
        )
        consumer_started, consumer_stdout_text, consumer_stderr_text = wait_for_consumer_start(
            consumer_process,
            stdout_path=consumer_stdout,
            stderr_path=consumer_stderr,
            timeout_s=consumer_start_timeout_s,
        )
        diagnostic["consumer_started"] = consumer_started
        diagnostic["consumer_exit_code"] = consumer_process.poll()
        if not consumer_started:
            row["status"] = "consumer_start_failed"
            diagnostic["status"] = "consumer_start_failed"
            diagnostic["consumer_stdout_tail"] = truncate_text(consumer_stdout_text)
            diagnostic["consumer_stderr_tail"] = truncate_text(consumer_stderr_text)
            return row, diagnostic
        if consumer_warmup_s > 0:
            time.sleep(consumer_warmup_s)

        producer_process = start_logged_process(
            spark_submit,
            producer_args,
            env=env,
            stdout_path=producer_stdout,
            stderr_path=producer_stderr,
        )
        try:
            producer_process.wait(timeout=producer_timeout_s)
        except subprocess.TimeoutExpired:
            row["status"] = "producer_timeout"
            diagnostic["status"] = "producer_timeout"
            diagnostic["producer_timeout_s"] = producer_timeout_s
            return row, diagnostic

        producer_exit_code = producer_process.poll()
        producer_completed = test_producer_completion(producer_stdout)
        diagnostic["producer_exit_code"] = producer_exit_code
        diagnostic["producer_completed"] = producer_completed
        diagnostic["producer_stdout_tail"] = truncate_text(read_log_text(producer_stdout))
        diagnostic["producer_stderr_tail"] = truncate_text(read_log_text(producer_stderr))
        if not producer_completed:
            row["status"] = "producer_incomplete"
            diagnostic["status"] = "producer_incomplete"
            return row, diagnostic

        time.sleep(post_run_wait_s)
        summary = summarize_live_metrics_file(metrics_path, bootstrap_servers)
        diagnostic["status"] = summary.get("status", "unknown")
        row.update(summary)
        row["metrics_status"] = summary.get("status", "unknown")
        if producer_exit_code not in (None, 0):
            row["producer_exit_code"] = producer_exit_code
        row["status"] = "ok" if summary.get("status") == "available" else str(summary.get("status", "unknown"))
        return row, diagnostic
    finally:
        stop_process_tree(producer_process)
        stop_process_tree(consumer_process)


def build_distributed_full_replay_validation_matrix() -> tuple[Dict[str, object], List[Dict[str, object]], List[Dict[str, object]]]:
    if not env_flag("ENABLE_DISTRIBUTED_VALIDATION", False):
        return (
            {
                "status": "disabled",
                "mode": "full_e2e_live_replay",
                "reason": "ENABLE_DISTRIBUTED_VALIDATION was not set.",
            },
            [],
            [],
        )

    spark_submit = resolve_spark_submit()
    runtime_jar = resolve_runtime_jar()
    if spark_submit is None:
        return (
            {
                "status": "unavailable",
                "mode": "full_e2e_live_replay",
                "reason": "spark-submit is unavailable, so full distributed replay cannot run.",
            },
            [],
            [],
        )
    if runtime_jar is None:
        return (
            {
                "status": "unavailable",
                "mode": "full_e2e_live_replay",
                "reason": "Runtime jar could not be resolved. Run the live suite once or set EFFITRACK_RUNTIME_JAR.",
            },
            [],
            [],
        )

    profiles = load_distributed_validation_profiles()
    if not profiles:
        return (
            {
                "status": "empty",
                "mode": "full_e2e_live_replay",
                "reason": "No distributed validation profiles were resolved from the environment.",
            },
            [],
            [],
        )

    spark_packages = os.getenv("SPARK_PACKAGES", "org.apache.spark:spark-streaming-kafka-0-10_2.11:2.3.1")
    kafka_topics_cmd = resolve_kafka_topics_command()
    run_token = time.strftime("%Y%m%d%H%M%S")
    rows: List[Dict[str, object]] = []
    diagnostics: List[Dict[str, object]] = []
    for profile in profiles:
        row, diagnostic = run_distributed_live_replay_profile(
            profile,
            runtime_jar=runtime_jar,
            spark_submit=spark_submit,
            spark_packages=spark_packages,
            kafka_topics_cmd=kafka_topics_cmd,
            run_token=run_token,
        )
        rows.append(row)
        diagnostics.append(diagnostic)

    success_count = sum(1 for row in rows if row.get("status") == "ok")
    broker_ready_count = sum(1 for row in rows if row.get("broker_reachable"))
    validation_scope = classify_validation_scope(profiles)
    summary = {
        "status": "available",
        "mode": "full_e2e_live_replay",
        "validation_scope": validation_scope,
        "profile_count": len(rows),
        "successful_profile_count": success_count,
        "failed_profile_count": len(rows) - success_count,
        "broker_reachable_profile_count": broker_ready_count,
        "spark_master_count": len({str(row["spark_master"]) for row in rows}),
        "kafka_profile_count": len({str(row["kafka_bootstrap_servers"]) for row in rows}),
        "runtime_jar": str(runtime_jar),
        "kafka_topic_admin_available": kafka_topics_cmd is not None,
        "note": distributed_validation_note("full_e2e_live_replay", validation_scope),
    }
    return summary, rows, diagnostics


def build_distributed_validation_matrix() -> tuple[Dict[str, object], List[Dict[str, object]], List[Dict[str, object]]]:
    mode = os.getenv("DISTRIBUTED_VALIDATION_MODE", "full_e2e_live_replay").strip() or "full_e2e_live_replay"
    if mode == "spark_probe_plus_broker_reachability":
        return build_distributed_probe_validation_matrix()
    if mode == "full_e2e_live_replay":
        return build_distributed_full_replay_validation_matrix()
    return (
        {
            "status": "invalid_mode",
            "mode": mode,
            "reason": "Unsupported DISTRIBUTED_VALIDATION_MODE. Use full_e2e_live_replay or spark_probe_plus_broker_reachability.",
        },
        [],
        [],
    )


def build_paper_alignment(use_direct_spark_metrics: bool) -> List[Dict[str, str]]:
    if use_direct_spark_metrics:
        return [
            {
                "section": "4.4.1 / Table 9",
                "parameter_grid_match": "Yes",
                "metric_alignment": "Partial",
                "status": "paper-equivalent",
                "note": "Data-flow sweep uses the same six flow-rate values and reports tasks/time/avg input/avg process columns.",
            },
            {
                "section": "4.4.2 / Tables 10-11",
                "parameter_grid_match": "Yes",
                "metric_alignment": "Strong",
                "status": "table10-equivalent table11-direct",
                "note": "Batch-interval sweep matches the paper grid. Table 10 task/batch/job outputs and Table 11 Spark JVM/storage peaks are now collected from a direct Spark benchmark job.",
            },
            {
                "section": "4.4.3 / Tables 12-13",
                "parameter_grid_match": "Yes",
                "metric_alignment": "Partial",
                "status": "paper-equivalent",
                "note": "The same flow-rate x batch-interval matrix is evaluated. Avg input/avg process and a delay matrix are produced with current runnable replay semantics.",
            },
            {
                "section": "4.4.4 / Table 14",
                "parameter_grid_match": "Yes",
                "metric_alignment": "Strong",
                "status": "core-sweep-direct",
                "note": "The core-count grid matches the paper and Spark JVM/storage peaks are collected from direct Spark runs at each core level.",
            },
            {
                "section": "4.4.5 / Table 15",
                "parameter_grid_match": "Yes",
                "metric_alignment": "Strong",
                "status": "partition-sweep-direct",
                "note": "The partition grid matches the paper and Spark storage/JVM/task metrics are collected from direct Spark runs.",
            },
        ]

    return [
        {
            "section": "4.4.1 / Table 9",
            "parameter_grid_match": "Yes",
            "metric_alignment": "Partial",
            "status": "paper-equivalent",
            "note": "Data-flow sweep uses the same six flow-rate values and reports tasks/time/avg input/avg process columns.",
        },
        {
            "section": "4.4.2 / Tables 10-11",
            "parameter_grid_match": "Yes",
            "metric_alignment": "Partial",
            "status": "table10-equivalent table11-analog",
            "note": "Batch-interval sweep matches the paper grid. Task/batch/job outputs align with Table 10; memory outputs remain Python analogs, not Spark JVM metrics.",
        },
        {
            "section": "4.4.3 / Tables 12-13",
            "parameter_grid_match": "Yes",
            "metric_alignment": "Partial",
            "status": "paper-equivalent",
            "note": "The same flow-rate x batch-interval matrix is evaluated. Avg input/avg process and a delay matrix are produced with current runnable replay semantics.",
        },
        {
            "section": "4.4.4 / Table 14",
            "parameter_grid_match": "Yes",
            "metric_alignment": "Analog only",
            "status": "core-sweep-analog",
            "note": "The core-count grid matches the paper, but memory columns are Python analogs rather than Spark JVM/storage counters.",
        },
        {
            "section": "4.4.5 / Table 15",
            "parameter_grid_match": "Yes",
            "metric_alignment": "Partial analog",
            "status": "partition-sweep-partial",
            "note": "The partition grid and throughput columns align. Memory columns are still analogs instead of direct Spark JVM/storage metrics.",
        },
    ]


def build_unavailable_routing_summary(reason: str) -> Dict[str, str]:
    return {
        "status": "unavailable",
        "reason": reason,
    }


def build_suite_results(
    benchmark_reservoir: Sequence[MessageSample],
    routing_reservoir: Sequence[MessageSample] | None,
    routing_score_source: Dict[str, object],
) -> Dict[str, object]:
    if routing_reservoir:
        policy_benchmark = build_policy_benchmark(routing_reservoir)
        routing_quality: Dict[str, object] = next(
            row for row in policy_benchmark if row["policy_name"] == "current_drop_policy"
        )
        threshold_sweep, threshold_recommendation = build_threshold_sweep(routing_reservoir)
    else:
        policy_benchmark = []
        routing_quality = build_unavailable_routing_summary(str(routing_score_source.get("reason", "Routing score source unavailable.")))
        threshold_sweep = []
        threshold_recommendation = {
            "status": "unavailable",
            "reason": str(routing_score_source.get("reason", "Routing score source unavailable.")),
        }
    repo_evidence = extract_repo_evidence()
    live_metrics = summarize_live_metrics()
    use_direct_spark_metrics = direct_spark_metrics_enabled()
    spark_probe_diagnostics: List[Dict[str, object]] = []
    distributed_validation_summary, distributed_validation_matrix, distributed_probe_diagnostics = (
        build_distributed_validation_matrix()
    )

    data_flow_results = [
        run_stream_simulation(
            benchmark_reservoir,
            data_flow_rate_ms=rate,
            batch_interval_s=10,
            duration_s=30,
            cores=4,
            partitions=1,
            work_rounds=10,
        )
        for rate in (10, 50, 100, 300, 400, 500)
    ]

    batch_interval_results = [
        run_stream_simulation(
            benchmark_reservoir,
            data_flow_rate_ms=500,
            batch_interval_s=interval,
            duration_s=100,
            cores=4,
            partitions=1,
            work_rounds=10,
        )
        for interval in (5, 10, 15, 20)
    ]
    if use_direct_spark_metrics:
        batch_interval_results = [
            overlay_direct_metrics(
                row,
                run_spark_resource_benchmark(
                    data_flow_rate_ms=500,
                    batch_interval_s=int(row["batch_interval_s"]),
                    duration_s=100,
                    cores=4,
                    partitions=1,
                    diagnostics=spark_probe_diagnostics,
                ),
            )
            for row in batch_interval_results
        ]

    delay_results = [
        run_stream_simulation(
            benchmark_reservoir,
            data_flow_rate_ms=flow,
            batch_interval_s=interval,
            duration_s=100,
            cores=4,
            partitions=1,
            work_rounds=10,
        )
        for flow in (500, 1000, 5000)
        for interval in (5, 10, 15, 20)
    ]

    core_results = [
        run_stream_simulation(
            benchmark_reservoir,
            data_flow_rate_ms=500,
            batch_interval_s=10,
            duration_s=60,
            cores=core_count,
            partitions=1,
            work_rounds=10,
        )
        for core_count in (1, 2, 4, 16)
    ]
    if use_direct_spark_metrics:
        core_results = [
            overlay_direct_metrics(
                row,
                run_spark_resource_benchmark(
                    data_flow_rate_ms=500,
                    batch_interval_s=10,
                    duration_s=60,
                    cores=int(row["cores"]),
                    partitions=1,
                    diagnostics=spark_probe_diagnostics,
                ),
            )
            for row in core_results
        ]

    partition_results = [
        run_stream_simulation(
            benchmark_reservoir,
            data_flow_rate_ms=500,
            batch_interval_s=10,
            duration_s=60,
            cores=16,
            partitions=partition_count,
            work_rounds=10,
        )
        for partition_count in (1, 3)
    ]
    if use_direct_spark_metrics:
        partition_results = [
            overlay_direct_metrics(
                row,
                run_spark_resource_benchmark(
                    data_flow_rate_ms=500,
                    batch_interval_s=10,
                    duration_s=60,
                    cores=16,
                    partitions=int(row["partitions"]),
                    diagnostics=spark_probe_diagnostics,
                ),
            )
            for row in partition_results
        ]

    effective_direct_spark_metrics = (
        use_direct_spark_metrics
        and all("storage_memory_kib" in row for row in batch_interval_results)
        and all("storage_memory_kib" in row for row in core_results)
        and all("storage_memory_kib" in row for row in partition_results)
    )

    comparison_rows = []
    for test in PAPER_TESTS:
        comparison_rows.append(
            {
                "id": test["id"],
                "section": test["section"],
                "name": test["name"],
                "variables": test["variables"],
                "metrics": test["metrics"],
                "repo_before": test["repo_before"],
                "added_now": "Yes",
                "single_command": r"ModelandPerformanceAnalysis\scripts\run\run_live_bv_kafka_suite.cmd",
                "evidence": test["evidence"],
            }
        )

    return {
        "pdf_reference": str(PDF_PATH),
        "repo_evidence": repo_evidence,
        "routing_score_source": routing_score_source,
        "comparison": comparison_rows,
        "paper_alignment": build_paper_alignment(effective_direct_spark_metrics),
        "routing_quality": routing_quality,
        "policy_benchmark": policy_benchmark,
        "threshold_sweep": threshold_sweep,
        "threshold_recommendation": threshold_recommendation,
        "live_metrics": live_metrics,
        "spark_probe_diagnostics": spark_probe_diagnostics,
        "distributed_validation_summary": distributed_validation_summary,
        "distributed_validation_matrix": distributed_validation_matrix,
        "distributed_probe_diagnostics": distributed_probe_diagnostics,
        "data_flow_rate_sweep": data_flow_results,
        "batch_interval_sweep": batch_interval_results,
        "delay_matrix": delay_results,
        "core_sweep": core_results,
        "partition_sweep": partition_results,
    }


def write_json_report(results: Dict[str, object]) -> None:
    RESULTS_DIR.mkdir(parents=True, exist_ok=True)
    JSON_PATH.write_text(json.dumps(results, indent=2), encoding="utf-8")


def write_workbook(results: Dict[str, object]) -> Path | None:
    if Workbook is None:
        return None

    RESULTS_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparison"
    ws.append(["Test_ID", "Paper_Section", "Paper_Test", "Variables", "Metrics", "Repo_Before", "Added_Now", "Command", "Evidence"])
    for row in results["comparison"]:
        ws.append([
            row["id"],
            row["section"],
            row["name"],
            row["variables"],
            row["metrics"],
            row["repo_before"],
            row["added_now"],
            row["single_command"],
            row["evidence"],
        ])

    append_key_value_sheet(wb, "RepoEvidence", results["repo_evidence"])
    append_key_value_sheet(wb, "RoutingScoreSource", results["routing_score_source"])
    append_sheet(wb, "PaperAlignment", results["paper_alignment"])
    append_key_value_sheet(wb, "RoutingQuality", results["routing_quality"])
    append_sheet(wb, "PolicyBenchmark", results["policy_benchmark"])
    append_key_value_sheet(wb, "ThresholdRecommendation", results["threshold_recommendation"])
    append_sheet(wb, "ThresholdSweep", results["threshold_sweep"])
    append_key_value_sheet(wb, "LiveMetrics", results["live_metrics"])
    append_sheet(wb, "SparkProbeDiag", results["spark_probe_diagnostics"])
    append_key_value_sheet(wb, "DistributedValidationSummary", results["distributed_validation_summary"])
    append_sheet(wb, "DistributedValidation", results["distributed_validation_matrix"])
    append_sheet(wb, "DistributedValidationDiag", results["distributed_probe_diagnostics"])
    append_sheet(wb, "DataFlowRate", results["data_flow_rate_sweep"])
    append_sheet(wb, "BatchInterval", results["batch_interval_sweep"])
    append_sheet(wb, "DelayMatrix", results["delay_matrix"])
    append_sheet(wb, "CoreSweep", results["core_sweep"])
    append_sheet(wb, "PartitionSweep", results["partition_sweep"])
    try:
        wb.save(WORKBOOK_PATH)
        return WORKBOOK_PATH
    except PermissionError:
        fallback_path = WORKBOOK_PATH.with_name(
            f"{WORKBOOK_PATH.stem}_{time.strftime('%Y%m%d-%H%M%S')}{WORKBOOK_PATH.suffix}"
        )
        wb.save(fallback_path)
        return fallback_path


def append_key_value_sheet(wb: Workbook, title: str, mapping: Dict[str, object]) -> None:
    ws = wb.create_sheet(title)
    ws.append(["Metric", "Value"])
    for key, value in mapping.items():
        ws.append([key, value])


def append_sheet(wb: Workbook, title: str, rows: Sequence[Dict[str, object]]) -> None:
    ws = wb.create_sheet(title)
    if not rows:
        ws.append(["empty"])
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])


def print_table(title: str, rows: Sequence[Dict[str, object]], columns: Sequence[str]) -> None:
    print(f"\n=== {title} ===")
    if not rows:
        print("(empty)")
        return
    widths: Dict[str, int] = {}
    for column in columns:
        values = [format_value(row.get(column, "")) for row in rows]
        widths[column] = max(len(column), *(len(value) for value in values))
    header = " | ".join(column.ljust(widths[column]) for column in columns)
    print(header)
    print("-+-".join("-" * widths[column] for column in columns))
    for row in rows:
        print(" | ".join(format_value(row.get(column, "")).ljust(widths[column]) for column in columns))


def format_value(value: object) -> str:
    if isinstance(value, float):
        if abs(value) >= 100:
            return f"{value:.2f}"
        return f"{value:.4f}"
    return str(value)


def print_delay_matrix(rows: Sequence[Dict[str, object]]) -> None:
    print("\n=== Table 13 Equivalent - Delay Matrix (mean scheduling delay ms) ===")
    grouped: Dict[int, Dict[int, float]] = {}
    for row in rows:
        grouped.setdefault(int(row["data_flow_rate_ms"]), {})[int(row["batch_interval_s"])] = float(row["mean_delay_ms"])

    batch_intervals = [5, 10, 15, 20]
    headers = ["flow_ms"] + [f"batch_{interval}s" for interval in batch_intervals]
    matrix_rows = []
    for flow in sorted(grouped):
        row = {"flow_ms": flow}
        for interval in batch_intervals:
            row[f"batch_{interval}s"] = grouped[flow].get(interval, 0.0)
        matrix_rows.append(row)
    print_table("Delay Matrix", matrix_rows, headers)


def main() -> int:
    if not NORMAL_RAW_PATH.exists():
        raise FileNotFoundError(f"Missing raw normal dataset: {NORMAL_RAW_PATH}")
    if not ANOMALOUS_RAW_PATH.exists():
        raise FileNotFoundError(f"Missing raw anomalous dataset: {ANOMALOUS_RAW_PATH}")

    use_prediction_scores = os.getenv("USE_PREDICTIONS_FOR_ROUTING", "0") == "1"
    if use_prediction_scores and PREDICTIONS_PATH.exists():
        routing_reservoir = load_prediction_samples(limit=None)
        routing_score_source = {
            "status": "ok",
            "source_type": "predictions_csv_legacy",
            "path": str(PREDICTIONS_PATH),
            "sample_count": len(routing_reservoir),
            "reason": "Routing analysis used the legacy predictions.csv score file because USE_PREDICTIONS_FOR_ROUTING=1 was set.",
        }
    else:
        routing_reservoir, routing_score_source = load_rf_score_samples(limit=None)
    benchmark_reservoir = load_live_replay_samples(limit=None)
    results = build_suite_results(benchmark_reservoir, routing_reservoir, routing_score_source)
    write_json_report(results)
    workbook_output_path = write_workbook(results)

    repo_evidence = results["repo_evidence"]
    using_direct_spark_metrics = any("direct" in row["status"] for row in results["paper_alignment"])
    failed_probe_count = sum(
        1
        for row in results["spark_probe_diagnostics"]
        if row.get("status") not in ("ok",)
    )
    print("BV big-data suite completed.")
    if using_direct_spark_metrics:
        print("Note: Table 11, Table 14, and Table 15 Spark JVM/storage metrics are collected from direct Spark benchmark runs.")
    else:
        print("Note: paper-side JVM/storage metrics are reproduced here as runnable current-repo replay benchmarks with Python memory analogs where direct Spark JVM counters are unavailable.")
    if results["spark_probe_diagnostics"]:
        print(f"Spark probe runs: {len(results['spark_probe_diagnostics'])}, failures: {failed_probe_count}")
    distributed_summary = results["distributed_validation_summary"]
    if distributed_summary.get("status") == "available":
        scope = str(distributed_summary.get("validation_scope", "distributed_cluster_replay"))
        prefix = "Distributed validation profiles"
        if scope == "local_multi_profile_replay":
            prefix = "Local multi-profile replay rows"
        elif scope == "hybrid_profile_matrix":
            prefix = "Hybrid validation profiles"
        successful_count = distributed_summary.get("successful_profile_count", distributed_summary.get("successful_probe_count", ""))
        print(
            f"{prefix}: "
            f"{distributed_summary['profile_count']}, "
            f"successful profiles: {successful_count}, "
            f"broker-ready profiles: {distributed_summary['broker_reachable_profile_count']}"
        )
    elif distributed_summary.get("status") not in ("disabled",):
        print(f"Distributed validation: {distributed_summary.get('status')} ({distributed_summary.get('reason', 'no detail')})")
    print(f"PDF reference: {results['pdf_reference']}")
    print(f"Benchmark source (raw replay): {NORMAL_RAW_PATH} + {ANOMALOUS_RAW_PATH}")
    if results["routing_score_source"].get("status") == "ok":
        routing_source_path = results["routing_score_source"].get("path") or results["routing_score_source"].get("output_path")
        print(f"Optional routing-score source: {routing_source_path}")
    else:
        print(f"Optional routing-score source: unavailable ({results['routing_score_source'].get('reason', 'unknown reason')})")
    print(f"Spark consumer batch interval: {repo_evidence['spark_batch_interval_s']} s")
    print(f"Consumer input topic: {repo_evidence['input_topic']}")
    print(
        f"Routing thresholds: anomaly>{repo_evidence['anomaly_threshold']}, normal<={repo_evidence['normal_threshold']}"
    )
    print(
        f"Routing topics: anomaly={repo_evidence['anomaly_topic']}, normal={repo_evidence['normal_topic']}, uncertain={repo_evidence['uncertain_topic']}"
    )
    print(f"Routing policy in code: {repo_evidence['routing_policy']}")
    print(f"Live metrics log: {repo_evidence['live_metrics_path']}")
    print(f"JSON report: {JSON_PATH}")
    if workbook_output_path is not None:
        print(f"Workbook: {workbook_output_path}")

    print_table(
        "Paper vs Repo Coverage",
        results["comparison"],
        ["id", "section", "name", "repo_before", "added_now"],
    )
    print_table(
        "Paper Alignment",
        results["paper_alignment"],
        ["section", "parameter_grid_match", "metric_alignment", "status", "note"],
    )
    if routing_reservoir:
        print_table(
            "Routing Quality (Optional Current Drop Policy)",
            [{"metric": key, "value": value} for key, value in results["routing_quality"].items()],
            ["metric", "value"],
        )
        print_table(
            "Policy Benchmark (Optional)",
            results["policy_benchmark"],
            [
                "policy_name",
                "anomaly_topic_count",
                "normal_topic_count",
                "uncertain_band_count",
                "review_topic_count",
                "precision_on_auto_alert",
                "recall_on_auto_alert",
                "effective_recall_with_review",
                "review_load_ratio",
                "fpr_on_auto_alert",
            ],
        )
        print_table(
            "Threshold Recommendation (Operational)",
            [{"metric": key, "value": value} for key, value in results["threshold_recommendation"].items()],
            ["metric", "value"],
        )
        print_table(
            "Threshold Sweep (Operational RF Score Replay)",
            results["threshold_sweep"],
            [
                "anomaly_threshold",
                "normal_threshold",
                "review_load_ratio",
                "uncertain_band_ratio",
                "precision_on_auto_alert",
                "recall_on_auto_alert",
                "effective_recall_with_review",
                "fpr_on_auto_alert",
                "delta_review_load_ratio_vs_baseline",
                "delta_effective_recall_with_review_vs_baseline",
                "is_baseline",
                "is_recommended",
            ],
        )
    else:
        print_table(
            "Routing Quality (Optional)",
            [{"metric": key, "value": value} for key, value in results["routing_quality"].items()],
            ["metric", "value"],
        )
    print_table(
        "Live Kafka Metrics",
        [{"metric": key, "value": value} for key, value in results["live_metrics"].items()],
        ["metric", "value"],
    )
    print_table(
        "Distributed Validation Summary",
        [{"metric": key, "value": value} for key, value in results["distributed_validation_summary"].items()],
        ["metric", "value"],
    )
    if results["distributed_validation_matrix"]:
        print_table(
            "Distributed Validation Matrix",
            results["distributed_validation_matrix"],
            [
                "validation_label",
                "spark_master",
                "kafka_bootstrap_servers",
                "broker_reachable",
                "data_flow_rate_ms",
                "batch_interval_s",
                "cores",
                "partitions",
                "status",
                "records_processed",
                "mean_e2e_latency_ms",
                "mean_total_consumer_lag",
                "mean_batch_duration_ms",
            ],
        )
    if results["distributed_probe_diagnostics"]:
        print_table(
            "Distributed Validation Diagnostics",
            results["distributed_probe_diagnostics"],
            [
                "validation_label",
                "spark_master",
                "kafka_bootstrap_servers",
                "status",
                "consumer_started",
                "consumer_exit_code",
                "producer_completed",
                "producer_exit_code",
            ],
        )
    if results["spark_probe_diagnostics"]:
        print_table(
            "Spark Probe Diagnostics",
            results["spark_probe_diagnostics"],
            [
                "data_flow_rate_ms",
                "batch_interval_s",
                "cores",
                "partitions",
                "status",
                "returncode",
                "output_json_exists",
            ],
        )
    print_table(
        "Table 9 Equivalent - Data Flow Rate Sweep",
        results["data_flow_rate_sweep"],
        ["data_flow_rate_ms", "tasks", "time_s", "avg_input_s", "avg_process_s"],
    )
    print_table(
        "Table 10 Equivalent - Batch Interval Sweep",
        results["batch_interval_sweep"],
        ["batch_interval_s", "tasks", "batch_count", "jobs"],
    )
    print_table(
        "Table 11 - Batch Memory",
        results["batch_interval_sweep"],
        ["batch_interval_s", "storage_memory_kib", "peak_jvm_on_heap_mib", "peak_jvm_off_heap_mib", "peak_storage_on_heap_mib"],
    )
    print_table(
        "Table 12 Equivalent - Avg Input / Avg Process",
        results["delay_matrix"],
        ["data_flow_rate_ms", "batch_interval_s", "avg_input_s", "avg_process_s"],
    )
    print_delay_matrix(results["delay_matrix"])
    print_table(
        "Table 14 - Core Sweep",
        results["core_sweep"],
        ["cores", "storage_memory_kib", "peak_jvm_on_heap_mib", "peak_jvm_off_heap_mib", "peak_storage_on_heap_mib"],
    )
    print_table(
        "Table 15 - Partition Sweep",
        results["partition_sweep"],
        ["partitions", "cores", "storage_memory_kib", "peak_jvm_on_heap_mib", "tasks", "task_time_s", "time_s", "avg_input_s", "avg_process_s"],
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
