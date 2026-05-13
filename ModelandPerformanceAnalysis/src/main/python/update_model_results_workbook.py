from __future__ import annotations

import csv
from collections import OrderedDict
from datetime import date
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


REPO_ROOT = Path(__file__).resolve().parents[4]
WORKBOOK_PATH = REPO_ROOT / "ModelResults.xlsx"
SUMMARY_PATH = REPO_ROOT / "ModelandPerformanceAnalysis" / "results_under_final" / "summary_all.csv"

HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
TITLE_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
FAMILY_FILLS = {
    "classical_ml": PatternFill(fill_type="solid", fgColor="E2F0D9"),
    "deep_learning": PatternFill(fill_type="solid", fgColor="FFF2CC"),
    "transformer": PatternFill(fill_type="solid", fgColor="FCE4D6"),
    "hybrid": PatternFill(fill_type="solid", fgColor="E4DFEC"),
}

SHEET_TITLES = OrderedDict(
    [
        ("FINAL_OVERVIEW", None),
        ("FINAL_UNDERSAMPLE", "hrss_undersample_optimized"),
    ]
)

DISPLAY_COLUMNS = [
    ("dataset", "Dataset"),
    ("model", "Model"),
    ("family", "Aile"),
    ("evaluation_strategy", "Degerlendirme"),
    ("accuracy_mean", "Accuracy Ort."),
    ("accuracy_std", "Accuracy Std."),
    ("precision_mean", "Precision Ort."),
    ("precision_std", "Precision Std."),
    ("recall_mean", "Recall Ort."),
    ("recall_std", "Recall Std."),
    ("f1_mean", "F1 Ort."),
    ("f1_std", "F1 Std."),
    ("baseline_f1_mean", "Random Baseline F1"),
    ("p_value", "p-degeri"),
    ("significant", "Anlamli mi"),
    ("direction", "Yon"),
    ("learning_rate", "Learning Rate"),
    ("batch_size", "Batch Size"),
    ("epochs", "Epoch"),
    ("optimizer", "Optimizer"),
    ("loss_function", "Loss"),
    ("hardware", "Donanim"),
]


def load_rows(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8") as handle:
        return list(csv.DictReader(handle))


def prettify_dataset(name: str) -> str:
    mapping = {
        "hrss_undersample_optimized": "HRSS Undersample Optimized",
    }
    return mapping.get(name, name)


def prettify_model(name: str) -> str:
    mapping = {
        "logistic_regression": "Logistic Regression",
        "random_forest": "Random Forest",
        "svm": "SVM",
        "decision_tree": "Decision Tree",
        "knn": "KNN",
        "naive_bayes": "Naive Bayes",
        "cnn": "CNN",
        "rnn": "RNN",
        "lstm": "LSTM",
        "gru": "GRU",
        "autoencoder": "Autoencoder",
        "vanilla_transformer": "Vanilla Transformer",
        "encoder_decoder_transformer": "Encoder-Decoder Transformer",
        "temporal_fusion_transformer": "Temporal Fusion Transformer",
        "cnn_lstm_hybrid": "CNN-LSTM Hybrid",
    }
    return mapping.get(name, name)


def prettify_family(name: str) -> str:
    mapping = {
        "classical_ml": "Classical ML",
        "deep_learning": "Deep Learning",
        "transformer": "Transformer",
        "hybrid": "Hybrid",
    }
    return mapping.get(name, name)


def prettify_direction(name: str) -> str:
    mapping = {
        "better_than_random": "Random baseline'dan daha iyi",
        "worse_than_random": "Random baseline'dan daha kotu",
        "equal_to_random": "Random baseline ile benzer",
    }
    return mapping.get(name, name)


def normalize_value(key: str, value: str):
    if key == "dataset":
        return prettify_dataset(value)
    if key == "model":
        return prettify_model(value)
    if key == "family":
        return prettify_family(value)
    if key == "direction":
        return prettify_direction(value)
    if key == "significant":
        return "Evet" if value == "True" else "Hayir"
    if value in {"N/A", ""}:
        return value
    try:
        return float(value)
    except ValueError:
        return value


def write_sheet(ws, rows: list[dict[str, str]], subtitle: str) -> None:
    total_columns = len(DISPLAY_COLUMNS)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)
    ws["A1"] = subtitle
    ws["A1"].font = Font(bold=True, size=13)
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_columns)
    ws["A2"] = (
        "Kaynak: ModelandPerformanceAnalysis/results/experiments/results_under_final/summary_all.csv | "
        "Metodoloji: classical ML = 5-fold CV, digerleri = hold-out + 3 run ortalamasi"
    )
    ws["A2"].alignment = Alignment(horizontal="center")

    for column_index, (_, header) in enumerate(DISPLAY_COLUMNS, start=1):
        cell = ws.cell(row=4, column=column_index, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    data_row = 5
    for row in rows:
        family_name = row["family"]
        fill = FAMILY_FILLS.get(family_name)
        for column_index, (key, _) in enumerate(DISPLAY_COLUMNS, start=1):
            value = normalize_value(key, row[key])
            cell = ws.cell(row=data_row, column=column_index, value=value)
            cell.alignment = Alignment(horizontal="center")
            if fill is not None:
                cell.fill = fill
            if isinstance(value, float):
                cell.number_format = "0.0000"
        data_row += 1

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = "A4:{}{}".format(get_column_letter(total_columns), max(4, data_row - 1))

    for column_index, (_, header) in enumerate(DISPLAY_COLUMNS, start=1):
        max_len = len(header)
        for row_index in range(5, data_row):
            value = ws.cell(row=row_index, column=column_index).value
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))
        ws.column_dimensions[get_column_letter(column_index)].width = min(max_len + 2, 28)


def main() -> None:
    rows = load_rows(SUMMARY_PATH)
    workbook = load_workbook(WORKBOOK_PATH)

    for sheet_name in SHEET_TITLES:
        if sheet_name in workbook.sheetnames:
            del workbook[sheet_name]

    for sheet_name, dataset_filter in SHEET_TITLES.items():
        filtered_rows = rows
        if dataset_filter is not None:
            filtered_rows = [row for row in rows if row["dataset"] == dataset_filter]

        worksheet = workbook.create_sheet(title=sheet_name)
        if dataset_filter is None:
            subtitle = "Final tum model sonuclari ({})".format(date.today().isoformat())
        else:
            subtitle = "{} sonuclari ({})".format(prettify_dataset(dataset_filter), date.today().isoformat())
        write_sheet(worksheet, filtered_rows, subtitle)

    workbook.save(WORKBOOK_PATH)


if __name__ == "__main__":
    main()
