from __future__ import annotations

import argparse
import csv
import sys
from pathlib import Path
from typing import Dict, List

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ModuleNotFoundError as exc:
    raise SystemExit(
        "openpyxl is required for workbook export. "
        "Install it with the active interpreter, for example: "
        "python -m pip install openpyxl"
    ) from exc


SCRIPT_ROOT = Path(__file__).resolve().parent
if str(SCRIPT_ROOT) not in sys.path:
    sys.path.insert(0, str(SCRIPT_ROOT))

from effitrack_eval.config import DATASET_REGISTRY
from effitrack_eval.data import find_repo_root


TITLE_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
DATA_FILL = PatternFill(fill_type="solid", fgColor="F8FBFF")

DEFAULT_DATASETS = [
    "hrss_anomalous_optimized",
    "hrss_undersample_optimized",
    "hrss_smote_optimized",
    "hrss_anomalous_standard",
    "hrss_undersample_standard",
    "hrss_smote_standard",
]

TURKISH_DATASET_LABELS = {
    "hrss_anomalous_optimized": "Optimized - Dengesiz",
    "hrss_undersample_optimized": "Optimized - Undersample Dengeli",
    "hrss_smote_optimized": "Optimized - SMOTE Dengeli",
    "hrss_anomalous_standard": "Standard - Dengesiz",
    "hrss_undersample_standard": "Standard - Undersample Dengeli",
    "hrss_smote_standard": "Standard - SMOTE Dengeli",
}

TURKISH_DESCRIPTIONS = {
    "hrss_anomalous_optimized": "Optimized ana veri setinin dengesiz hali; normal ve anomali sinif oranlari korunur.",
    "hrss_undersample_optimized": "Optimized veri setinde cogunluk sinifi olan normal kayitlar azaltilarak sinif dengesi kurulmustur.",
    "hrss_smote_optimized": "Optimized veri setinde azinlik sinifi olan anomali kayitlari SMOTE ile artirilip sinif dengesi kurulmustur.",
    "hrss_anomalous_standard": "Standard ana veri setinin dengesiz hali; normal ve anomali sinif oranlari korunur.",
    "hrss_undersample_standard": "Standard veri setinde cogunluk sinifi olan normal kayitlar azaltilarak sinif dengesi kurulmustur.",
    "hrss_smote_standard": "Standard veri setinde azinlik sinifi olan anomali kayitlari SMOTE ile artirilip sinif dengesi kurulmustur.",
}

LOCALE_SETTINGS = {
    "en": {
        "title": "Selected Real Dataset Summary",
        "sheet_name": "DATASETS",
        "columns": {
            "dataset": "dataset",
            "display_name": "display_name",
            "description": "description",
            "total_rows": "total_rows",
            "normal_rows": "normal_rows",
            "anomaly_rows": "anomaly_rows",
            "label_column": "label_column",
            "relative_path": "relative_path",
        },
    },
    "tr": {
        "title": "Secili Gercek Veri Seti Ozeti",
        "sheet_name": "VERI_SETLERI",
        "columns": {
            "dataset": "veri_seti_kodu",
            "display_name": "veri_seti_adi",
            "description": "aciklama",
            "total_rows": "toplam_satir",
            "normal_rows": "normal_satir",
            "anomaly_rows": "anomali_satir",
            "label_column": "etiket_kolonu",
            "relative_path": "dosya_yolu",
        },
    },
}


def _resolve_dataset_path(dataset_name: str) -> Path:
    repo_root = find_repo_root()
    relative_path = DATASET_REGISTRY[dataset_name].relative_path
    return (repo_root / relative_path).resolve()


def _count_labels(dataset_path: Path) -> Dict[str, int | str]:
    with dataset_path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        total_rows = 0
        label_column = None
        label_counts: Dict[str, int] = {}

        for row in reader:
            if label_column is None:
                for candidate in ("Labels", "label", "Label", "class", "target", "y"):
                    if candidate in row:
                        label_column = candidate
                        break
            total_rows += 1
            if label_column is None:
                continue
            label_value = str(row[label_column]).strip()
            label_counts[label_value] = label_counts.get(label_value, 0) + 1

    return {
        "total_rows": total_rows,
        "label_column": label_column or "N/A",
        "normal_rows": int(label_counts.get("0", 0) + label_counts.get("0.0", 0)),
        "anomaly_rows": int(label_counts.get("1", 0) + label_counts.get("1.0", 0)),
    }


def _build_rows(dataset_names: List[str], locale: str) -> List[Dict[str, object]]:
    rows: List[Dict[str, object]] = []
    repo_root = find_repo_root()
    use_turkish = locale == "tr"

    for dataset_name in dataset_names:
        dataset_spec = DATASET_REGISTRY[dataset_name]
        dataset_path = _resolve_dataset_path(dataset_name)
        counts = _count_labels(dataset_path)
        description = dataset_spec.description
        display_name = dataset_name
        if use_turkish:
            description = TURKISH_DESCRIPTIONS.get(dataset_name, description)
            display_name = TURKISH_DATASET_LABELS.get(dataset_name, dataset_name)

        rows.append(
            {
                "dataset": dataset_name,
                "display_name": display_name,
                "description": description,
                "total_rows": counts["total_rows"],
                "normal_rows": counts["normal_rows"],
                "anomaly_rows": counts["anomaly_rows"],
                "label_column": counts["label_column"],
                "relative_path": str(dataset_path.relative_to(repo_root)),
            }
        )

    return rows


def _apply_layout(ws, title: str, columns: List[str], rows: List[Dict[str, object]]) -> None:
    total_columns = len(columns)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=13)
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = Alignment(horizontal="center")

    for column_index, column in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=column_index, value=column)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    row_index = 4
    for row in rows:
        for column_index, column in enumerate(columns, start=1):
            value = row[column]
            cell = ws.cell(row=row_index, column=column_index, value=value)
            cell.fill = DATA_FILL
            if isinstance(value, int):
                cell.number_format = "#,##0"
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        row_index += 1

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = "A3:{}{}".format(get_column_letter(total_columns), max(3, row_index - 1))

    for column_index, column in enumerate(columns, start=1):
        max_length = len(column)
        for data_row in range(4, row_index):
            value = ws.cell(data_row, column_index).value
            max_length = max(max_length, len(str(value)))
        ws.column_dimensions[get_column_letter(column_index)].width = min(max_length + 2, 48)


def main() -> int:
    parser = argparse.ArgumentParser(description="Export selected dataset overview workbook")
    parser.add_argument(
        "--datasets",
        nargs="*",
        default=DEFAULT_DATASETS,
        help="Dataset names to include in the overview workbook",
    )
    parser.add_argument(
        "--output",
        required=True,
        help="Target workbook path",
    )
    parser.add_argument(
        "--locale",
        choices=sorted(LOCALE_SETTINGS.keys()),
        default="en",
        help="Workbook language",
    )
    args = parser.parse_args()

    settings = LOCALE_SETTINGS[args.locale]
    rows = _build_rows(args.datasets, locale=args.locale)
    internal_columns = [
        "dataset",
        "display_name",
        "description",
        "total_rows",
        "normal_rows",
        "anomaly_rows",
        "label_column",
        "relative_path",
    ]
    column_labels = settings["columns"]

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = settings["sheet_name"]
    display_rows: List[Dict[str, object]] = []
    for row in rows:
        display_rows.append({column_labels[column]: row[column] for column in internal_columns})
    display_columns = [column_labels[column] for column in internal_columns]
    _apply_layout(worksheet, settings["title"], display_columns, display_rows)

    output_path = Path(args.output).resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    print(output_path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
