from __future__ import annotations

import csv
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


ROOT = Path(__file__).resolve().parents[2]
OUT_PATH = ROOT / "report" / "BV_Test_Plan_EffiTrack_TR.xlsx"

HEADER_CANDIDATES = [
    ROOT / "Data" / "HRSS_SMOTE_standard.csv",
    ROOT / "SparkKafkaStreaming" / "src" / "main" / "scala" / "data" / "HRSS_SMOTE_standard.csv",
]
PREDICTIONS_PATH = ROOT / "SparkKafkaStreaming" / "predictions.csv"

RAW_STANDARD_TOTAL = 23645
RAW_STANDARD_NORMAL = 17975
RAW_STANDARD_ANOMALY = 5670

RAW_OPTIMIZED_TOTAL = 19634
RAW_OPTIMIZED_NORMAL = 15117
RAW_OPTIMIZED_ANOMALY = 4517

SMOTE_K = 5
STANDARD_UNDERSAMPLE_FRACTION = RAW_STANDARD_ANOMALY / RAW_STANDARD_NORMAL
OPTIMIZED_UNDERSAMPLE_FRACTION = RAW_OPTIMIZED_ANOMALY / RAW_OPTIMIZED_NORMAL
STANDARD_SMOTE_GENERATED = RAW_STANDARD_NORMAL - RAW_STANDARD_ANOMALY
OPTIMIZED_SMOTE_GENERATED = RAW_OPTIMIZED_NORMAL - RAW_OPTIMIZED_ANOMALY

THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(vertical="top", wrap_text=True)

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
NOTE_FILL = PatternFill("solid", fgColor="F3F3F3")
GOOD_FILL = PatternFill("solid", fgColor="E2F0D9")
WARN_FILL = PatternFill("solid", fgColor="FCE4D6")


def pick_existing_path(paths: list[Path]) -> Path:
    for path in paths:
        if path.exists():
            return path
    raise FileNotFoundError("No candidate CSV header path exists.")


def load_header(path: Path) -> list[str]:
    with path.open(newline="", encoding="utf-8") as handle:
        return next(csv.reader(handle))


def summarize_predictions(path: Path) -> dict[str, float]:
    if not path.exists():
        return {
            "total": 0,
            "anomaly": 0,
            "normal": 0,
            "uncertain": 0,
            "anomaly_ratio": 0.0,
            "normal_ratio": 0.0,
            "uncertain_ratio": 0.0,
        }

    total = anomaly = normal = uncertain = 0
    with path.open(newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            total += 1
            score = float(row["predictions"])
            if score > 0.8:
                anomaly += 1
            elif score <= 0.2:
                normal += 1
            else:
                uncertain += 1

    return {
        "total": total,
        "anomaly": anomaly,
        "normal": normal,
        "uncertain": uncertain,
        "anomaly_ratio": anomaly / total if total else 0.0,
        "normal_ratio": normal / total if total else 0.0,
        "uncertain_ratio": uncertain / total if total else 0.0,
    }


def signal_meta(column: str) -> tuple[str, str, str]:
    if column == "Timestamp":
        return ("System", "time", "Kaydin zaman ekseni")
    if column == "Labels":
        return ("System", "label", "Sinif etiketi")

    parts = column.split("_")
    subsystem = parts[2]

    if column.startswith("I_"):
        return (subsystem, "current", f"{subsystem} icin giris akimi")
    if column.endswith("_power"):
        return (subsystem, "power", f"{subsystem} icin cikis gucu")
    return (subsystem, "voltage", f"{subsystem} icin cikis voltaji")


def add_rows(ws, rows: list[list[object]]) -> None:
    for row in rows:
        ws.append(row)


def style_sheet(
    ws,
    widths: dict[str, int] | None = None,
    input_cols: Iterable[int] = (),
    warning_cols: Iterable[int] = (),
) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP
        cell.border = BORDER

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = WRAP
            cell.border = BORDER

    for col_idx in input_cols:
        for row_idx in range(2, ws.max_row + 1):
            ws.cell(row=row_idx, column=col_idx).fill = INPUT_FILL

    for col_idx in warning_cols:
        for row_idx in range(2, ws.max_row + 1):
            ws.cell(row=row_idx, column=col_idx).fill = WARN_FILL

    if widths:
        for col_name, width in widths.items():
            ws.column_dimensions[col_name].width = width


def build_workbook() -> Workbook:
    header_path = pick_existing_path(HEADER_CANDIDATES)
    header = load_header(header_path)
    predictions = summarize_predictions(PREDICTIONS_PATH)

    total_columns = len(header)
    feature_columns = [col for col in header if col not in {"Timestamp", "Labels"}]
    feature_count = len(feature_columns)
    subsystems = sorted({signal_meta(col)[0] for col in feature_columns})
    subsystem_count = len(subsystems)

    wb = Workbook()

    # Sheet 1
    ws = wb.active
    ws.title = "Kullanim_Notu"
    note_rows = [
        ["Baslik", "EffiTrack buyuk veri ve akisa dayali BV test plani"],
        ["Amac", "Kafka + Spark + model serving akisinin hangi asamada hangi giris parametreleriyle test edilecegini Excel uzerinden netlestirmek."],
        ["Hazirlama mantigi", "Bu dosya mevcut EffiTrack kodu, makaledeki sayisal sonuclar ve WaterLog referans reposundaki operasyon adimlari uzerine kurulmustur."],
        ["WaterLog'tan alinan fikir", "WaterLog README; Kafka topic olusturma, producer-consumer zinciri, delay analizi icin realtimeListenersdelay ve Spark UI (4040) kullanimini vurguluyor."],
        ["EffiTrack'ta mevcut eksik", "Makale tarafinda end-to-end latency, Kafka publish latency, throughput, consumer lag, ariza toparlanma suresi ve mesaj kaybi olculmuyor."],
        ["EffiTrack'ta mevcut durum", "Canli BV akisi model-input uzerinden girer; cikis topic'leri anomalies3, normal_data ve uncertain_data olarak ayrilir."],
        ["EffiTrack'ta mevcut eksik", "0.2 < prediction <= 0.8 bandindaki kayitlar icin ozel bir routing yolu yok; bu blind zone test edilmeden BV sign-off verilmemeli."],
        ["Mevcut topoloji", "Spark ayari local[*]; mevcut repo gercek multi-node dagitik kurulum degil, tek host uzerinde cok cekirdekli calisma durumunda."],
        ["Nasil kullanilir", "Sari hucreler test sirasinda doldurulacak alanlardir. Gri hucreler referans not veya repo/makale tabanli baslangic degerleridir."],
        ["Ana sayfalar", "Adim_Bazli_Test_Akisi = hangi adimda ne verilecek; Yuk_ve_Durum_Senaryolari = load/fault testleri; KPI_ve_Log_Alanlari = ne loglanacak; Makale_Eksik_Olcumler = kapatilmasi gereken aciklar."],
        ["Kaynak 1", "WaterLog GitHub README: https://github.com/SedaBalta/WaterLog-Cybersecurity-aware-Log-Management-System-for-Water-Critical-Infrastructures"],
        ["Kaynak 2", "ScienceDirect abstract: https://www.sciencedirect.com/science/article/pii/S1568494624013875"],
    ]
    add_rows(ws, note_rows)
    ws["A1"].font = Font(size=14, bold=True)
    ws["B1"].font = Font(size=14, bold=True)
    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row=row_idx, column=1).fill = SECTION_FILL
        ws.cell(row=row_idx, column=1).font = Font(bold=True)
        ws.cell(row=row_idx, column=1).alignment = WRAP
        ws.cell(row=row_idx, column=2).alignment = WRAP
        ws.cell(row=row_idx, column=1).border = BORDER
        ws.cell(row=row_idx, column=2).border = BORDER
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 120

    # Sheet 2
    ws = wb.create_sheet("Adim_Bazli_Test_Akisi")
    ws.append(
        [
            "Adim_No",
            "Asama",
            "Bilesen",
            "Test_Konusu",
            "Girdi_Parametresi",
            "Ornek_Deger",
            "Verilecegi_Yer",
            "Beklenen_Cikti",
            "Olculecek_Metrik",
            "Kabul_Kriteri",
            "Olculen_Deger",
            "Durum",
            "Kaynak",
            "Not",
        ]
    )

    step_rows = [
        ["PRE-01", "Veri girisi", "Spark read", "CSV dosyasi okunuyor mu", "filePath", "Data/HRSS_anomalous_standard.csv", "DatasetAndPreprocessing Scala akisi", "Dataset yuklenmis DataFrame", "load_success", "Basarili yukleme, parse hatasi yok", "", "", "HRSSDataPreprocessing.scala", "Ilk smoke testi."],
        ["PRE-02", "Veri girisi", "Spark read", "Kolon semasi dogru mu", "header=true, inferSchema=true", "true, true", "spark.read.option(...)", f"Toplam {total_columns} kolon, model icin {feature_count} feature", "column_count, schema_match", f"Toplam kolon = {total_columns}; model feature = {feature_count}", "", "", "HRSSDataPreprocessing.scala; report/main.tex", "Timestamp + Labels + 18 sensor ozelligi."],
        ["PRE-03", "Veri kalitesi", "Spark preprocessing", "Null/eksik deger taramasi", "null_count_per_column", "Tum kolonlar", "Dataset yuklendikten hemen sonra", "Kolon bazli null raporu", "null_count_per_column", "0 veya null-handling kurali dokumante edilmeli", "", "", "HRSSDataPreprocessing.scala", "Makalede null count sonucu yok."],
        ["PRE-04", "Veri kalitesi", "Raw standard veri", "Sinif dagilimi standard raw", "normal_count, anomaly_count", f"{RAW_STANDARD_NORMAL}, {RAW_STANDARD_ANOMALY}", "groupBy(Labels).count()", "Raw standard class dagilimi", "class_ratio", f"Toplam = {RAW_STANDARD_TOTAL}; anomaly rate = {RAW_STANDARD_ANOMALY / RAW_STANDARD_TOTAL:.4f}", "", "", "report/main.tex", "Karsilastirma baseline'i."],
        ["PRE-05", "Veri kalitesi", "Raw optimized veri", "Sinif dagilimi optimized raw", "normal_count, anomaly_count", f"{RAW_OPTIMIZED_NORMAL}, {RAW_OPTIMIZED_ANOMALY}", "groupBy(Labels).count()", "Raw optimized class dagilimi", "class_ratio", f"Toplam = {RAW_OPTIMIZED_TOTAL}; anomaly rate = {RAW_OPTIMIZED_ANOMALY / RAW_OPTIMIZED_TOTAL:.4f}", "", "", "report/main.tex", "Karsilastirma baseline'i."],
        ["BAL-01", "Dengeleme", "Undersampling", "Standard veri undersampling", "samplingFraction", f"{STANDARD_UNDERSAMPLE_FRACTION:.4f}", "majorityDF.sample(...)", "Siniflar esitlenmis cikti", "balanced_class_counts", "Normal = Anomaly = 5670", "", "", "Undersampling.scala; report/main.tex", "Standard raw veri icin."],
        ["BAL-02", "Dengeleme", "Undersampling", "Optimized veri undersampling", "samplingFraction", f"{OPTIMIZED_UNDERSAMPLE_FRACTION:.4f}", "majorityDF.sample(...)", "Siniflar esitlenmis cikti", "balanced_class_counts", "Normal = Anomaly = 4517", "", "", "report/main.tex", "Kod ornegi standard veri icin ama optimized icin de ayni mantik test edilmeli."],
        ["BAL-03", "Dengeleme", "SMOTE", "Standard veri SMOTE", "k, numToGenerate", f"{SMOTE_K}, {STANDARD_SMOTE_GENERATED}", "SMOTE.scala", "SMOTE dengeli veri seti", "balanced_class_counts, synthetic_row_count", f"Normal = Anomaly = {RAW_STANDARD_NORMAL}", "", "", "SMOTE.scala; report/main.tex", "k = 5 kodda sabit."],
        ["BAL-04", "Dengeleme", "SMOTE", "Optimized veri SMOTE", "k, numToGenerate", f"{SMOTE_K}, {OPTIMIZED_SMOTE_GENERATED}", "SMOTE.scala mantigi", "SMOTE dengeli veri seti", "balanced_class_counts, synthetic_row_count", f"Normal = Anomaly = {RAW_OPTIMIZED_NORMAL}", "", "", "report/main.tex", "Optimized veri icin de ayni parametre seti kayda gecmeli."],
        ["KFK-01", "Akis altyapisi", "Kafka", "Broker erisimi", "bootstrap.servers", "localhost:9092", "Kafka producer/consumer config", "Broker ayakta ve erisilebilir", "broker_reachability", "Baglanti hatasi olmamali", "", "", "KafkaConsumerRF.scala; KafkaProducerRF.scala", "Su an lokal broker varsayiliyor."],
        ["KFK-02", "Akis altyapisi", "Kafka", "Input topic kontrati", "topic", "model-input", "Consumer topic subscription", "Input topic'ten veri okunur", "topic_read_success", "Topic tanimi sabit ve dokumante olmali", "", "", "KafkaConsumerRF.scala", "BV sirasinda ayni isim kullanilmali."],
        ["KFK-03", "Akis altyapisi", "Kafka", "Consumer ayarlari", "group.id, auto.offset.reset, enable.auto.commit", "rf-group, latest, false", "Kafka params", "Tutarlı consumer davranisi", "consumer_lag, offset_progress", "Lag test sirasinda izlenebilir olmali", "", "", "KafkaConsumerRF.scala", "Offset davranisi test edilmeden performans yorumu zayif kalir."],
        ["STR-01", "Akis altyapisi", "Spark Streaming", "Micro-batch araligi", "batch_interval_sec", "10", "new StreamingContext(conf, Seconds(10))", "Her 10 saniyede bir batch", "effective_batch_interval_sec", "<= 10 sn sapma ile calisma", "", "", "KafkaConsumerRF.scala", "WaterLog benzeri delay testi icin kritik."],
        ["INF-01", "Model serving", "RF pipeline artifact", "Kaydedilmis Spark ML pipeline erisimi", "rf_model_path", "SparkKafkaStreaming/model_artifacts/random_forest_smote_standard_pipeline", "KafkaConsumerRF.scala icindeki PipelineModel.load", "Skor ureten yuklenmis model", "model_load_success, model_load_time_ms", "Model acilisinda hatasiz yuklenmeli", "", "", "KafkaConsumerRF.scala; RandomForestModelExport.scala", "HTTP inference bagimliligi artik yok."],
        ["INF-02", "Model serving", "RF egitim/export akisi", "Model export argumanlari", "RF_MODEL_PATH, RF_TRAINING_DATA_PATH", "random_forest_smote_standard_pipeline, HRSS_SMOTE_standard.csv", "RandomForestModelExport.scala", "Kaydedilmis pipeline olusur", "model_artifact_created", "Artifact olusmali ve bos olmamali", "", "", "RandomForestModelExport.scala", "Bu akis SMOTE standard ile egitilmis RF pipeline'ini uretir."],
        ["RTE-01", "Routing", "Anomaly topic", "Anomali routing esigi", "prediction_threshold", "> 0.8", "predictions.filter('predictions > 0.8')", "Anomaly topic'e yazim", "anomaly_routed_count", "Esige uyan tum kayitlar anomalies3'e gitmeli", "", "", "KafkaAnomalyDetection.scala", "Makaledeki operational KPI ile baglanti kurulacak."],
        ["RTE-02", "Routing", "Normal topic", "Normal routing esigi", "prediction_threshold", "<= 0.2", "predictions.filter('predictions <= 0.2')", "Normal topic'e yazim", "normal_routed_count", "Esige uyan tum kayitlar normal_data'ya gitmeli", "", "", "KafkaAnomalyDetection.scala", "Normal-state davranisi icin."],
        ["RTE-03", "Routing", "Blind zone", "Belirsiz skor bandi handling", "prediction_band", "0.2 < p <= 0.8", "Routing kurallari arasinda acikta kaliyor", "Ya ozel topic ya da acik drop kurali", "uncertain_rate, dropped_uncertain_count", "Kuralsiz kayit kalmamali veya bilincli kabul edilmeli", "", "", "KafkaAnomalyDetection.scala; predictions.csv", f"Mevcut sample'da {predictions['uncertain']} kayit, oran {predictions['uncertain_ratio']:.2%}."],
        ["RTE-04", "Routing", "Topic kontrati", "Topic isimleri tutarli mi", "input/output/alert topic isimleri", "model-input, anomalies3, normal_data, uncertain_data", "Tum Scala akislari birlikte", "Tek bir end-to-end topic haritasi", "topic_contract_match", "BV oncesi tek topic matrisi onaylanmali", "", "", "KafkaConsumerRF.scala; KafkaProducerRF.scala; KafkaAnomalyDetection.scala", "Raw-data producer dogrudan model-input topic'ine yazar."],
        ["VAL-01", "Dogrulama", "Offline accuracy", "Raw standard performans tekrar testi", "F1, precision, recall, FPR", "0.9630, 0.9635, 0.9626, 0.0115", "Offline replay / ground truth ile skorla", "Makaledeki baseline tekrar edilir", "F1, precision, recall, normal_fpr", "F1 >= 0.95 ve FPR <= 0.015", "", "", "report/main.tex", "Best raw model = Random Forest."],
        ["VAL-02", "Dogrulama", "Offline accuracy", "Raw optimized performans tekrar testi", "F1, precision, recall, FPR", "0.9539, 0.9579, 0.9500, 0.0125", "Offline replay / ground truth ile skorla", "Makaledeki baseline tekrar edilir", "F1, precision, recall, normal_fpr", "F1 >= 0.95 ve FPR <= 0.015", "", "", "report/main.tex", "Iki davranis da yeniden dogrulanmali."],
        ["PERF-01", "Performans", "Latency", "Event-to-score gecikmesi", "event_ts, prediction_end_ts", "ornek: 1712345678, 1712345689", "Consumer + predictor loglari", "ms bazli detection latency", "detection_latency_ms", "BV kabul esigi ekip tarafindan netlestirilmeli", "", "", "Makale ve kod icinde dogrudan yok", "Makaledeki en buyuk eksiklerden biri."],
        ["PERF-02", "Performans", "Latency", "Score-to-Kafka gecikmesi", "prediction_end_ts, kafka_send_ts", "ornek: 1712345689, 1712345691", "KafkaAnomalyDetection publish noktasi", "ms bazli publish latency", "kafka_publish_latency_ms", "BV kabul esigi ekip tarafindan netlestirilmeli", "", "", "Makale ve kod icinde dogrudan yok", "WaterLog delay mantigina daha yakin metrik."],
        ["PERF-03", "Performans", "Kapasite", "Throughput ve lag testi", "processed_count, elapsed_sec, broker_offset, consumer_offset", "ornek: 5000, 60, 120000, 119950", "Spark UI + Kafka lag izlencesi", "records/s ve consumer lag", "throughput_records_s, consumer_lag", "Nominal yukte lag sifira yakin olmali", "", "", "Spark 4040 fikri WaterLog README'den", "Makale tarafinda yok."],
        ["PERF-04", "Performans", "Kaynak kullanimi", "CPU/RAM/GC izleme", "cpu_pct, ram_gb", "ornek: 62%, 9.8 GB", "OS monitor + Spark UI", "Kaynak tuketim profili", "cpu_pct, ram_gb", "Uzun sureli testte kontrol disi artis olmamali", "", "", "Makale ve kod icinde raporlanmiyor", "Long-run test icin gerekli."],
        ["RES-01", "Dayaniklilik", "Ariza senaryosu", "RF model artifact yuklenemezse ne olur", "fault_window_sec", "60", "Model artifact path'ini bozarak consumer baslat", "Hata logu + toparlanma davranisi", "error_rate, recovery_time_sec", "Silent drop olmamali, toparlanma suresi olculmeli", "", "", "KafkaConsumerRF.scala; RandomForestModelExport.scala", "Kritik zayif nokta artik model artifact erisimi."],
        ["RES-02", "Dayaniklilik", "Ariza senaryosu", "Kafka broker restart", "restart_minute", "5. dakika", "Load testi sirasinda broker restart", "Akis devam / toparlanma", "recovery_time_sec, message_loss_rate", "Mesaj kaybi dokumante edilmeli ve toparlanma olculmeli", "", "", "Kafka akisi", "Makaledeki eksik operational test."],
        ["RES-03", "Dayaniklilik", "Schema drift", "Gerekli kolonlardan biri eksik gelirse", "missing_column", "I_w_HR_Weg", "Producer payload / CSV header", "Kayit reject veya quarantine edilmeli", "schema_error_rate", "Sebepsiz crash olmamali", "", "", "Model input semasi", "18 feature dogrulugu korunmali."],
        ["RES-04", "Dayaniklilik", "Veri bozulmasi", "Null row / NaN row enjeksiyonu", "null_ratio", "1%, 5%, 10%", "Producer payload", "Invalid kayit handling", "null_error_rate, dropped_record_rate", "Kuralsiz yutma olmamali", "", "", "HRSSDataPreprocessing.scala", "Makale null handling'i raporlamiyor."],
        ["RES-05", "Dayaniklilik", "Mesaj kalitesi", "Duplicate mesaj testi", "duplicate_ratio", "5%", "Producer tarafinda ayni message_id tekrar gonder", "Duplicate davranisi gorulur", "duplicate_detected_count", "En azindan duplicate oraninin logu tutulmali", "", "", "Kafka producer-consumer zinciri", "Idempotency su an tanimli degil."],
        ["RES-06", "Dayaniklilik", "Zaman sirasi", "Out-of-order timestamp testi", "out_of_order_ratio", "10%", "Timestamp alanini bozarak replay", "Latency hesaplari bozulmadan calismali", "ordering_error_rate", "Hesaplanan KPI'lar nan/invalid olmamali", "", "", "Timestamp kullanan tum yorumlar", "Process-level analiz icin onemli."],
    ]
    add_rows(ws, step_rows)
    style_sheet(
        ws,
        widths={
            "A": 12,
            "B": 14,
            "C": 16,
            "D": 24,
            "E": 22,
            "F": 22,
            "G": 24,
            "H": 22,
            "I": 22,
            "J": 22,
            "K": 16,
            "L": 12,
            "M": 30,
            "N": 36,
        },
        input_cols=(11, 12),
        warning_cols=(14,),
    )
    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row=row_idx, column=6).fill = NOTE_FILL
        ws.cell(row=row_idx, column=13).fill = NOTE_FILL

    # Sheet 3
    ws = wb.create_sheet("Yuk_ve_Durum_Senaryolari")
    ws.append(
        [
            "Senaryo_ID",
            "Senaryo_Tipi",
            "Amac",
            "Girdi_Hizi_records_s",
            "Sure_dk",
            "Batch_s",
            "Partition_Sayisi",
            "Replication_Factor",
            "Input_Topic",
            "Ariza_Enjeksiyonu",
            "Takip_Edilecek_KPIlar",
            "Ornek_Kabul",
            "Olculen_Deger",
            "Durum",
            "Not",
        ]
    )
    load_rows = [
        ["SCN-01", "Smoke", "Aklin dogru acildigini gormek", 10, 2, 10, 1, 1, "model-input", "Yok", "load_success, topic_read_success, dropped_record_rate", "Crash yok, drop yok", "", "", "Ilk calistirma testi."],
        ["SCN-02", "Nominal", "Temel operasyon yuku", 50, 10, 10, 1, 1, "model-input", "Yok", "detection_latency_ms, kafka_publish_latency_ms, consumer_lag", "Lag sifira yakin, crash yok", "", "", "Laptop/local kurulum icin mantikli baslangic."],
        ["SCN-03", "Medium", "Yuk arttiginda davranis", 100, 10, 10, 3, 1, "model-input", "Yok", "throughput_records_s, consumer_lag, cpu_pct, ram_gb", "Backlog kontrol altinda kalmali", "", "", "Partition etkisi gorulmeye baslar."],
        ["SCN-04", "Stress", "Ust sinir davranisi", 250, 10, 10, 3, 1, "model-input", "Yok", "consumer_lag, dropped_record_rate, recovery_time_sec", "Kalici birikme olmamali veya raporlanmali", "", "", "Single-host siniri burada gorulebilir."],
        ["SCN-05", "Burst", "Ani patlama yukunde tepki", 500, 2, 10, 3, 1, "model-input", "30 sn burst", "peak_lag, time_to_recover_batches, dropped_record_rate", "Burst sonrasi en gec 3 batch icinde toparlama", "", "", "WaterLog delay fikrine yakin test."],
        ["SCN-06", "Long-run", "Uzun sureli stabilite", 100, 60, 10, 3, 1, "model-input", "Yok", "cpu_pct, ram_gb, error_rate, consumer_lag", "Memory leak belirtisi olmamali", "", "", "Makale tarafinda eksik."],
        ["SCN-07", "Partition scale", "Partition artisina tepki", 100, 10, 10, 6, 1, "model-input", "Yok", "throughput_records_s, consumer_lag", "3 partition'a gore iyilesme veya en azindan bozulmama", "", "", "Tek broker oldugu icin replication 1 kalir."],
        ["SCN-08", "Fault", "Predict servisi kesilirse", 50, 10, 10, 1, 1, "model-input", "3.-4. dk arasi predict API down", "error_rate, recovery_time_sec, message_loss_rate", "Silent loss olmamali, toparlanma loglanmali", "", "", "HTTP bagimliligi testi."],
        ["SCN-09", "Fault", "Broker restart olursa", 50, 10, 10, 1, 1, "model-input", "5. dakikada Kafka restart", "message_loss_rate, consumer_rejoin_sec, lag_after_restart", "Toparlanma suresi kaydedilmeli", "", "", "WaterLog tarafindaki operasyonel kurulum adimina dogrudan bagli."],
        ["SCN-10", "Fault", "Schema drift", 20, 5, 10, 1, 1, "model-input", "1 zorunlu feature eksik", "schema_error_rate, crash_count", "Crash olmamali, hata sebebi loglanmali", "", "", "18 feature uyumu korunmali."],
        ["SCN-11", "Fault", "Null / NaN enjeksiyonu", 20, 5, 10, 1, 1, "model-input", "5% null veya NaN", "null_error_rate, dropped_record_rate", "Kuralsiz yutma olmamali", "", "", "Preprocessing dayanimi."],
        ["SCN-12", "Fault", "Duplicate + order drift", 50, 5, 10, 1, 1, "model-input", "5% duplicate, 10% out-of-order timestamp", "duplicate_detected_count, ordering_error_rate", "KPI hesaplari bozulmamali", "", "", "Message quality testi."],
    ]
    add_rows(ws, load_rows)
    style_sheet(
        ws,
        widths={
            "A": 12,
            "B": 12,
            "C": 22,
            "D": 16,
            "E": 10,
            "F": 10,
            "G": 12,
            "H": 16,
            "I": 16,
            "J": 24,
            "K": 28,
            "L": 28,
            "M": 18,
            "N": 12,
            "O": 34,
        },
        input_cols=(13, 14),
        warning_cols=(15,),
    )

    # Sheet 4
    ws = wb.create_sheet("KPI_ve_Log_Alanlari")
    ws.append(
        [
            "KPI_Adi",
            "Formul_veya_Log_Alani",
            "Logda_Tutulacak_Degiskenler",
            "Ornek_Format",
            "Loglanacagi_Katman",
            "Neden_Gerekli",
            "Makalede_Eksik_mi",
            "Not",
        ]
    )
    kpi_rows = [
        ["detection_latency_ms", "prediction_end_ts - event_ts", "message_id, event_ts, prediction_end_ts", "142", "Consumer + predict servisi", "Event goruldugu anda skora ne kadar surede gidildigini gosterir.", "Evet", "Makaledeki en kritik eksik runtime KPI."],
        ["kafka_publish_latency_ms", "kafka_send_ts - prediction_end_ts", "message_id, prediction_end_ts, kafka_send_ts, output_topic", "18", "Kafka publish noktasi", "Skor sonrasi alert topic'e yazma gecikmesini gosterir.", "Evet", "WaterLog delay mantigina daha yakin."],
        ["throughput_records_s", "processed_records / elapsed_sec", "batch_id, processed_records, batch_start_ts, batch_end_ts", "257.4", "Spark Streaming", "Hat kapasitesi ve sizing icin gerekli.", "Evet", "Makalede yok."],
        ["consumer_lag", "broker_offset - consumer_offset", "topic, partition, broker_offset, consumer_offset", "35", "Kafka consumer", "Akin yetisip yetismedigini gosterir.", "Evet", "Nominal yukte sifira yakin olmali."],
        ["message_loss_rate", "lost_messages / input_messages", "message_id, input_count, delivered_count", "0.002", "Producer + consumer + output topic", "Silent drop var mi yok mu gosterir.", "Evet", "Broker restart ve predict down testlerinde zorunlu."],
        ["uncertain_rate", "uncertain_messages / total_scored", "message_id, prediction_score, route_label", "0.324", "Routing katmani", "0.2-0.8 bandindaki blind zone'u gorunur yapar.", "Evet", f"Mevcut predictions.csv icin {predictions['uncertain_ratio']:.2%}."],
        ["alert_delivery_success", "delivered_alerts / expected_alerts", "message_id, expected_topic, actual_topic", "0.998", "Kafka output kontrolu", "Anomaly alert'lerin kaybolup kaybolmadigini gosterir.", "Evet", "BV sign-off icin kritik."],
        ["recovery_time_sec", "first_ok_ts - fault_start_ts", "fault_start_ts, first_ok_ts, fault_type", "41", "Fault senaryolari", "Ariza sonrasi sistemin ne kadar hizli toparlandigini gosterir.", "Evet", "Predict down ve broker restart icin."],
        ["schema_error_rate", "schema_error_records / input_records", "message_id, schema_ok, error_type", "0.010", "Input validation", "Eksik kolon veya tip bozulmasi etkisini gosterir.", "Evet", "Schema drift testi icin."],
        ["null_error_rate", "null_invalid_records / input_records", "message_id, null_columns, error_type", "0.007", "Input validation", "Null / NaN satirlarina nasil davranildigini gosterir.", "Evet", "Makale null handling raporlamiyor."],
        ["duplicate_detected_count", "count(distinct duplicate_message_id)", "message_id, first_seen_ts, duplicate_flag", "27", "Kafka / stream logic", "Idempotency veya duplicate issue var mi gosterir.", "Evet", "Su an kodda yok."],
        ["cpu_pct", "OS monitor metric", "host, process_name, cpu_pct", "61.8", "Host monitor", "Long-run ve stress testte darbogazi gosterir.", "Evet", "Makalede sadece donanim var, runtime kaynak kullanimi yok."],
        ["ram_gb", "OS monitor metric", "host, process_name, ram_gb", "9.4", "Host monitor", "Bellek sizinti belirtisini gosterir.", "Evet", "Long-run icin gerekli."],
        ["time_to_first_alert_sec", "first_alert_ts - anomaly_start_ts", "anomaly_start_ts, first_alert_ts", "7.2", "Alert katmani", "Erken uyari kapasitesini gosterir.", "Evet", "Operasyonel anlamli bir KPI."],
    ]
    add_rows(ws, kpi_rows)
    style_sheet(
        ws,
        widths={
            "A": 22,
            "B": 26,
            "C": 34,
            "D": 14,
            "E": 22,
            "F": 30,
            "G": 14,
            "H": 34,
        },
    )
    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row=row_idx, column=8).fill = NOTE_FILL

    # Sheet 5
    ws = wb.create_sheet("Konfig_ve_Giris_Parametreleri")
    ws.append(
        [
            "Alan",
            "Parametre",
            "Mevcut_Deger",
            "Ornek_Test_Degeri",
            "Tip",
            "Kaynak",
            "Aciklama",
        ]
    )
    config_rows = [
        ["Calisma modu", "spark.master", "local[*]", "local[*] veya cluster URL", "mode", "KafkaAnomalyDetection.scala; SMOTE.scala; Undersampling.scala", "Mevcut repo tek host, cok cekirdek."],
        ["Surum", "Scala", "2.11.8", "2.11.8", "version", "build.sbt", "Kod tabaninin hedef surumu."],
        ["Surum", "Spark", "2.3.1", "2.3.1", "version", "build.sbt", "Streaming ve ML kutuphaneleri bu surume bagli."],
        ["Surum", "Kafka client", "2.0.0", "2.0.0", "version", "SparkKafkaStreaming/build.sbt", "Kafka istemci kutuphanesi."],
        ["Kafka", "bootstrap.servers", "localhost:9092", "localhost:9092 veya BV broker listesi", "endpoint", "Kafka Scala dosyalari", "Broker adresi."],
        ["Kafka", "consumer group", "rf-group", "rf-group", "group id", "KafkaConsumerRF.scala", "Lag ve offset takibi icin."],
        ["Kafka", "input topic", "model-input", "model-input", "topic", "KafkaConsumerRF.scala", "Model-serving input."],
        ["Kafka", "raw producer topic", "model-input", "model-input", "topic", "KafkaProducerRF.scala", "Producer consumer input topic'ine dogrudan yazar."],
        ["Kafka", "anomaly output topic", "anomalies3", "anomalies3", "topic", "KafkaAnomalyDetection.scala", "Anomali alarmlari icin."],
        ["Kafka", "normal output topic", "normal_data", "normal_data", "topic", "KafkaAnomalyDetection.scala", "Normal akisin yonlendirilmesi icin."],
        ["Streaming", "batch interval", "10", "10 / 5 / 2", "sec", "KafkaConsumerRF.scala", "Delay ve throughput testleri icin degistirilebilir."],
        ["Inference", "RF pipeline artifact", "SparkKafkaStreaming/model_artifacts/random_forest_smote_standard_pipeline", "ayni", "path", "KafkaConsumerRF.scala; RandomForestModelExport.scala", "Model artifact bagimliligi."],
        ["Scoring", "anomaly threshold", "> 0.8", "> 0.8", "rule", "KafkaAnomalyDetection.scala", "Alert routing kuralı."],
        ["Scoring", "normal threshold", "<= 0.2", "<= 0.2", "rule", "KafkaAnomalyDetection.scala", "Normal routing kuralı."],
        ["Scoring", "uncertain band", "0.2 < p <= 0.8", "ayni", "rule", "Derived", "Routing blind zone."],
        ["Preprocessing", "SMOTE k", str(SMOTE_K), str(SMOTE_K), "int", "SMOTE.scala", "En yakin komsu sayisi."],
        ["Preprocessing", "Standard undersample fraction", f"{STANDARD_UNDERSAMPLE_FRACTION:.4f}", f"{STANDARD_UNDERSAMPLE_FRACTION:.4f}", "ratio", "Undersampling.scala + makale sayilari", "Standard veri dengesi icin."],
        ["Preprocessing", "Optimized undersample fraction", f"{OPTIMIZED_UNDERSAMPLE_FRACTION:.4f}", f"{OPTIMIZED_UNDERSAMPLE_FRACTION:.4f}", "ratio", "Makale sayilari", "Optimized veri dengesi icin."],
        ["Preprocessing", "Standard SMOTE uretilecek satir", str(STANDARD_SMOTE_GENERATED), str(STANDARD_SMOTE_GENERATED), "rows", "Makale sayilari", "Standard veri icin sentetik satir sayisi."],
        ["Preprocessing", "Optimized SMOTE uretilecek satir", str(OPTIMIZED_SMOTE_GENERATED), str(OPTIMIZED_SMOTE_GENERATED), "rows", "Makale sayilari", "Optimized veri icin sentetik satir sayisi."],
        ["Veri", "Toplam kolon", str(total_columns), str(total_columns), "count", str(header_path.relative_to(ROOT)), "Timestamp + Labels + sensor kolonlari."],
        ["Veri", "Model feature sayisi", str(feature_count), str(feature_count), "count", str(header_path.relative_to(ROOT)), "Timestamp ve Labels haric feature adedi."],
        ["Veri", "Alt sistem sayisi", str(subsystem_count), str(subsystem_count), "count", str(header_path.relative_to(ROOT)), "BRU, BHR, BHL, BLO, HR, HL."],
        ["Predictions sample", "Toplam skorlanan kayit", str(predictions["total"]), str(predictions["total"]), "rows", "SparkKafkaStreaming/predictions.csv", "Mevcut sample boyutu."],
        ["Predictions sample", "Anomaly routed kayit", str(predictions["anomaly"]), str(predictions["anomaly"]), "rows", "SparkKafkaStreaming/predictions.csv", "p > 0.8."],
        ["Predictions sample", "Normal routed kayit", str(predictions["normal"]), str(predictions["normal"]), "rows", "SparkKafkaStreaming/predictions.csv", "p <= 0.2."],
        ["Predictions sample", "Uncertain kayit", str(predictions["uncertain"]), str(predictions["uncertain"]), "rows", "SparkKafkaStreaming/predictions.csv", "0.2 < p <= 0.8."],
    ]
    add_rows(ws, config_rows)
    style_sheet(
        ws,
        widths={
            "A": 18,
            "B": 24,
            "C": 20,
            "D": 20,
            "E": 12,
            "F": 32,
            "G": 36,
        },
    )
    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row=row_idx, column=3).fill = NOTE_FILL
        ws.cell(row=row_idx, column=4).fill = INPUT_FILL

    # Sheet 6
    ws = wb.create_sheet("Makale_Eksik_Olcumler")
    ws.append(
        [
            "Eksik_Alan",
            "Su_Anki_Durum",
            "Neden_Yetersiz",
            "Hangi_Testsayfasinda_Kapanir",
            "Olculmesi_Gereken_Degerler",
            "Kapanis_Kriteri",
            "Aciklama",
        ]
    )
    missing_rows = [
        ["End-to-end detection latency", "Makale raporlamiyor", "Gercek zaman iddiasi sayisal destek almiyor", "Adim_Bazli_Test_Akisi + Yuk_ve_Durum_Senaryolari", "event_ts, prediction_end_ts", "detection_latency_ms kayda alinmis olmali", "WaterLog'ta delay fikri var; bizde sayisal tabloya donusmeli."],
        ["Kafka publish latency", "Makale raporlamiyor", "Alert'in topic'e ne kadar gec yazildigi bilinmiyor", "Adim_Bazli_Test_Akisi", "prediction_end_ts, kafka_send_ts", "kafka_publish_latency_ms kayda alinmis olmali", "Alert dagitimi icin gerekli."],
        ["Throughput", "Makale raporlamiyor", "Sistem kapasitesi bilinmiyor", "Yuk_ve_Durum_Senaryolari", "processed_records, elapsed_sec", "records/s hesaplanmis olmali", "Nominal ve stress senaryolarinda ayri ayri."],
        ["Consumer lag", "Makale raporlamiyor", "Akin yetisip yetismedigi anlasilamiyor", "Yuk_ve_Durum_Senaryolari", "broker_offset, consumer_offset", "lag metrik olarak kayda alinmis olmali", "Spark UI tek basina yeterli degil, Excel'e yazilmali."],
        ["CPU / RAM kullanimi", "Makale yalnizca donanim veriyor", "Runtime darbogazlari gorunmuyor", "Yuk_ve_Durum_Senaryolari", "cpu_pct, ram_gb", "Ozellikle long-run ve stress'te izlenmis olmali", "Kaynak tuketimi BV icin gerekli."],
        ["Message loss", "Makale raporlamiyor", "Silent drop varsa sistem guvenilmez olur", "Yuk_ve_Durum_Senaryolari", "input_count, delivered_count", "message_loss_rate hesaplanmis olmali", "Broker restart ve predict down testleriyle kapanir."],
        ["Uncertain band handling", "Kodda acik ama makalede yok", "0.2-0.8 bandi belirsiz", "Adim_Bazli_Test_Akisi", "uncertain_count, uncertain_rate", "Acik routing veya kabul karari olmali", "Mevcut kod bosluk birakiyor."],
        ["Schema drift handling", "Makale raporlamiyor", "Saha verisi her zaman egitim semasina uymaz", "Yuk_ve_Durum_Senaryolari", "schema_error_rate, error_type", "Eksik kolon halinde crash olmamali", "BV icin kritik."],
        ["Null / NaN handling", "Makale raporlamiyor", "Eksik veri durumda akisin ne yaptigi belli degil", "Yuk_ve_Durum_Senaryolari", "null_error_rate, dropped_record_rate", "Null senaryosu sonuclandirilmali", "Preprocessing dayanimi."],
        ["Restart recovery", "Makale raporlamiyor", "Servis arizasi sonrasi toparlanma suresi yok", "Yuk_ve_Durum_Senaryolari", "fault_start_ts, first_ok_ts", "recovery_time_sec olculmeli", "Operasyonel dayanıklılık KPI'si."],
        ["Topic contract consistency", "Tek input + uc output", "End-to-end akis model-input -> anomalies3/normal_data/uncertain_data olarak netlesmis olmali", "Adim_Bazli_Test_Akisi", "topic_contract_match", "Tek bir topic matrisi onaylanmali", "model-input/anomalies3/normal_data/uncertain_data."],
        ["Real distribution level", "Spark local[*]", "Gercek multi-node test yapilmadan dagitik denmemeli", "Konfig_ve_Giris_Parametreleri", "node_count, broker_count, executor_count", "Topoloji acikca yazilmali", "Dogruluk acisindan gerekli."],
        ["Partition scaling impact", "Makale raporlamiyor", "Partition artisinin throughput/lag'a etkisi bilinmiyor", "Yuk_ve_Durum_Senaryolari", "partition_count, throughput, lag", "En az 1/3/6 partition karsilastirmasi olmali", "WaterLog topic kurulumundan esinlenildi."],
    ]
    add_rows(ws, missing_rows)
    style_sheet(
        ws,
        widths={
            "A": 24,
            "B": 20,
            "C": 26,
            "D": 28,
            "E": 28,
            "F": 24,
            "G": 34,
        },
        warning_cols=(7,),
    )
    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row=row_idx, column=2).fill = NOTE_FILL
        ws.cell(row=row_idx, column=5).fill = INPUT_FILL

    # Sheet 7
    ws = wb.create_sheet("WaterLog_Referans")
    ws.append(
        [
            "WaterLog_Unsuru",
            "Referans_Detay",
            "EffiTrack_Karsiligi",
            "Bizde_Test_Aksiyonu",
            "Kaynak",
        ]
    )
    waterlog_rows = [
        ["Kafka topic kurulum adimi", "README'de topic create komutu, partitions=1, replication-factor=1 ornegi veriliyor.", "Bizde topic isimleri var ama tek bir kurulum matrisi yok.", "1, 3 ve 6 partition icin test senaryosu ac.", "WaterLog README"],
        ["Producer-consumer zinciri", "README producer ve consumer'i ayri adimlar olarak tarif ediyor.", "Bizde KafkaProducerRF, KafkaConsumerRF ve KafkaAnomalyDetection birlikte RF tabanli akisi olusturuyor.", "End-to-end topic kontratini Excel'de tek satirda onayla.", "WaterLog README"],
        ["Delay analizi", "realtimeListenersdelay ile batch gecikmesi olculuyor.", "Bizde latency icin dogrudan log alani yok.", "detection_latency_ms ve kafka_publish_latency_ms loglarini ekle ve testte doldur.", "WaterLog README"],
        ["Spark 4040 performans kontrolu", "README Spark web port 4040 uzerinden performans testi diyor.", "Bizde de Spark UI kullanilabilir ama tabloya yazilmiyor.", "Spark UI gozlemini throughput, lag, CPU/RAM ile Excel'e indir.", "WaterLog README"],
        ["ELK / Kibana sink", "WaterLog Elasticsearch index ve Kibana dashboard kullaniyor.", "Bizde downstream sink Kafka topic ile bitiyor.", "Simdilik zorunlu degil; ama alert delivery success ve topic bazli log dogrulamasini yap.", "WaterLog README"],
        ["Iki topic uzerinden filtreleme", "myCustomFunc() ile iki farkli topic'e dagitimdan bahsediliyor.", "Bizde anomalies3 ve normal_data var, uncertain band yok.", "Ucuncu bandi test et veya bilincli drop karari yaz.", "WaterLog README"],
    ]
    add_rows(ws, waterlog_rows)
    style_sheet(
        ws,
        widths={"A": 22, "B": 34, "C": 28, "D": 34, "E": 18},
    )
    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row=row_idx, column=2).fill = NOTE_FILL
        ws.cell(row=row_idx, column=4).fill = INPUT_FILL

    # Sheet 8
    ws = wb.create_sheet("Sinyal_Semasi")
    ws.append(
        [
            "Kolon",
            "Alt_Sistem",
            "Sinyal_Tipi",
            "Modelde_Kullaniliyor_mu",
            "Process_Analizinde_Kullaniliyor_mu",
            "Aciklama",
        ]
    )
    for column in header:
        subsystem, signal_type, description = signal_meta(column)
        model_used = "Hayir" if column in {"Timestamp", "Labels"} else "Evet"
        process_used = "Evet"
        ws.append([column, subsystem, signal_type, model_used, process_used, description])
    style_sheet(
        ws,
        widths={"A": 24, "B": 12, "C": 14, "D": 18, "E": 22, "F": 38},
    )

    return wb


if __name__ == "__main__":
    workbook = build_workbook()
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUT_PATH)
    print(f"Wrote {OUT_PATH}")
