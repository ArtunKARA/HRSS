from __future__ import annotations

import argparse
import csv
from collections import OrderedDict, defaultdict
from pathlib import Path
from typing import Dict, Iterable, List, Optional

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ModuleNotFoundError as exc:
    raise SystemExit(
        "openpyxl is required for workbook export. "
        "Install it with the active interpreter, for example: "
        "python -m pip install openpyxl"
    ) from exc


HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
TITLE_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
FAMILY_FILLS = {
    "classical_ml": PatternFill(fill_type="solid", fgColor="E2F0D9"),
    "deep_learning": PatternFill(fill_type="solid", fgColor="FFF2CC"),
    "transformer": PatternFill(fill_type="solid", fgColor="FCE4D6"),
    "hybrid": PatternFill(fill_type="solid", fgColor="E4DFEC"),
}


def _read_csv_rows(path: Path) -> List[Dict[str, str]]:
    with path.open(newline="", encoding="utf-8") as handle:
        return list(csv.DictReader(handle))


def _normalize_value(value: Optional[str]):
    if value in (None, ""):
        return ""
    if value == "True":
        return "Evet"
    if value == "False":
        return "Hayir"
    try:
        return float(value)
    except ValueError:
        return value


def _collect_columns(rows: Iterable[Dict[str, str]]) -> List[str]:
    columns: List[str] = []
    preferred = [
        "dataset",
        "model",
        "family",
        "evaluation_strategy",
        "accuracy_mean",
        "accuracy_std",
        "precision_mean",
        "precision_std",
        "recall_mean",
        "recall_std",
        "f1_mean",
        "f1_std",
        "baseline_f1_mean",
        "baseline_f1_std",
        "p_value",
        "significant",
        "direction",
        "learning_rate",
        "batch_size",
        "epochs",
        "optimizer",
        "loss_function",
        "class_weight_mode",
        "tuning_profile",
        "hardware",
    ]

    seen = set()
    for column in preferred:
        for row in rows:
            if column in row and column not in seen:
                columns.append(column)
                seen.add(column)
                break

    for row in rows:
        for column in row.keys():
            if column not in seen:
                columns.append(column)
                seen.add(column)

    return columns


def _write_sheet(ws, title: str, rows: List[Dict[str, str]], columns: List[str]) -> None:
    total_columns = len(columns)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=13)
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = Alignment(horizontal="center")

    for index, column in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=index, value=column)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    row_index = 4
    for row in rows:
        family_fill = FAMILY_FILLS.get(str(row.get("family", "")))
        for column_index, column in enumerate(columns, start=1):
            value = _normalize_value(row.get(column))
            cell = ws.cell(row=row_index, column=column_index, value=value)
            cell.alignment = Alignment(horizontal="center")
            if family_fill is not None:
                cell.fill = family_fill
            if isinstance(value, float):
                cell.number_format = "0.0000"
        row_index += 1

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = "A3:{}{}".format(get_column_letter(total_columns), max(3, row_index - 1))

    for column_index, column in enumerate(columns, start=1):
        max_length = len(column)
        for data_row in range(4, row_index):
            value = ws.cell(data_row, column_index).value
            max_length = max(max_length, len(str(value)))
        ws.column_dimensions[get_column_letter(column_index)].width = min(max_length + 2, 28)


def main() -> int:
    parser = argparse.ArgumentParser(description="Export experiment summary CSV files to Excel")
    parser.add_argument("--summary", required=True, help="Path to summary.csv or summary_all.csv")
    parser.add_argument("--output", required=True, help="Target workbook path")
    parser.add_argument("--tuning-summary", default=None, help="Optional tuning_summary.csv path")
    args = parser.parse_args()

    summary_path = Path(args.summary).resolve()
    output_path = Path(args.output).resolve()
    tuning_summary_path = Path(args.tuning_summary).resolve() if args.tuning_summary else None

    summary_rows = _read_csv_rows(summary_path)
    workbook = Workbook()
    workbook.remove(workbook.active)

    columns = _collect_columns(summary_rows)
    grouped_rows: Dict[str, List[Dict[str, str]]] = defaultdict(list)
    for row in summary_rows:
        grouped_rows[str(row.get("dataset", "unknown"))].append(row)

    workbook.create_sheet("OVERVIEW")
    _write_sheet(workbook["OVERVIEW"], "All experiment results", summary_rows, columns)

    for dataset_name, rows in OrderedDict(sorted(grouped_rows.items())).items():
        sheet_name = dataset_name[:31]
        workbook.create_sheet(sheet_name)
        _write_sheet(workbook[sheet_name], "{} results".format(dataset_name), rows, columns)

    if tuning_summary_path and tuning_summary_path.exists():
        tuning_rows = _read_csv_rows(tuning_summary_path)
        tuning_columns = _collect_columns(tuning_rows)
        workbook.create_sheet("TUNING")
        _write_sheet(workbook["TUNING"], "DL tuning summary", tuning_rows, tuning_columns)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    print(output_path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
