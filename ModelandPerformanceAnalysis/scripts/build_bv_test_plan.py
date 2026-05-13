from __future__ import annotations

import csv
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
OUT_PATH = ROOT / "report" / "BV_Test_Plan_EffiTrack.xlsx"
DATA_HEADER_PATH = ROOT / "Data" / "HRSS_SMOTE_standard.csv"
PREDICTIONS_PATH = ROOT / "SparkKafkaStreaming" / "predictions.csv"

THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
NOTE_FILL = PatternFill("solid", fgColor="F3F3F3")
GOOD_FILL = PatternFill("solid", fgColor="E2F0D9")
WRAP = Alignment(vertical="top", wrap_text=True)


def load_header(path: Path) -> list[str]:
    with path.open(newline="", encoding="utf-8") as f:
        return next(csv.reader(f))


def summarize_predictions(path: Path) -> dict[str, float]:
    total = anomaly = normal = uncertain = 0
    with path.open(newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            total += 1
            p = float(row["predictions"])
            if p > 0.8:
                anomaly += 1
            elif p <= 0.2:
                normal += 1
            else:
                uncertain += 1
    return {
        "total": total,
        "anomaly": anomaly,
        "normal": normal,
        "uncertain": uncertain,
        "anomaly_ratio": anomaly / total if total else 0.0,
        "normal_ratio": normal / total if total else 0.0,
        "uncertain_ratio": uncertain / total if total else 0.0,
    }


def signal_meta(column: str) -> tuple[str, str, str]:
    if column == "Timestamp":
        return ("System", "time", "Observed process timeline")
    if column == "Labels":
        return ("System", "label", "Supervision label")

    parts = column.split("_")
    subsystem = parts[2]
    if column.startswith("I_"):
        signal_type = "current"
        desc = f"Input current for {subsystem}"
    elif column.endswith("_power"):
        signal_type = "power"
        desc = f"Output power for {subsystem}"
    else:
        signal_type = "voltage"
        desc = f"Output voltage for {subsystem}"
    return (subsystem, signal_type, desc)


def style_sheet(ws, widths: dict[str, int] | None = None, input_cols: Iterable[int] = ()) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP
        cell.border = BORDER

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = WRAP
            cell.border = BORDER

    for col_idx in input_cols:
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=col_idx).fill = INPUT_FILL

    if widths:
        for col, width in widths.items():
            ws.column_dimensions[col].width = width


def add_rows(ws, rows: list[list[object]]) -> None:
    for row in rows:
        ws.append(row)


def build_workbook() -> Workbook:
    header = load_header(DATA_HEADER_PATH)
    predictions = summarize_predictions(PREDICTIONS_PATH)

    feature_cols = [c for c in header if c not in {"Timestamp", "Labels"}]
    subsystems = sorted({signal_meta(c)[0] for c in feature_cols})
    current_count = sum(1 for c in feature_cols if signal_meta(c)[1] == "current")
    power_count = sum(1 for c in feature_cols if signal_meta(c)[1] == "power")
    voltage_count = sum(1 for c in feature_cols if signal_meta(c)[1] == "voltage")

    wb = Workbook()
    ws = wb.active
    ws.title = "README"

    readme_rows = [
        ["EffiTrack BV Test Workbook", ""],
        ["Purpose", "Turn the study results and the current Kafka-Spark implementation into an executable BV-side test plan."],
        ["How to use", "Fill only the yellow cells during BV execution. Gray cells are current baselines or notes extracted from the study/codebase."],
        ["Key interpretation", "The current repo is not cluster-distributed yet. Spark is configured in local[*], so the current state is single-host, multi-core execution."],
        ["Key interpretation", "Predictions between 0.2 and 0.8 are not routed to a dedicated Kafka topic in the current implementation. This must be explicitly tested and resolved."],
        ["Key interpretation", "The live BV path now uses a single producer-to-consumer contract on model-input, with anomalies3 / normal_data / uncertain_data as output topics."],
        ["Main sheets", "BV_Test_Cases = operational checklist; Offline_Baselines = study metrics; Signal_Coverage = expected signals; Config_Map = runtime/config values; Metric_Definitions = KPI meaning."],
        ["Suggested usage", "Use this workbook as the single source for BV sign-off: data availability, signal coverage, routing logic, detection quality, and runtime behaviour."],
    ]
    add_rows(ws, readme_rows)
    ws["A1"].font = Font(size=14, bold=True)
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=1).fill = SECTION_FILL
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=1).border = BORDER
        ws.cell(row=row, column=2).border = BORDER
        ws.cell(row=row, column=1).alignment = WRAP
        ws.cell(row=row, column=2).alignment = WRAP
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 120

    # Test cases
    ws = wb.create_sheet("BV_Test_Cases")
    headers = [
        "Test_ID",
        "Area",
        "Metric",
        "Why_It_Matters",
        "Baseline_or_Config_Value",
        "Unit",
        "Suggested_BV_Acceptance",
        "BV_Measured_Value",
        "Measurement_Method",
        "Source",
        "Priority",
        "Status",
        "Notes",
    ]
    ws.append(headers)

    test_rows = [
        ["ARCH-01", "Architecture", "Spark execution mode", "Shows whether execution is truly distributed or only local multi-core.", "local[*]", "mode", "Cluster definition documented or local-mode limitation accepted", "", "Inspect SparkConf / runtime submission", "SparkKafkaStreaming/src/main/scala/spark/KafkaAnomalyDetection.scala; KafkaConsumerRF.scala", "High", "", "Current repo is single-host, not cluster-distributed."],
        ["ARCH-02", "Architecture", "Kafka bootstrap servers", "Defines broker topology and whether messaging is local or distributed.", "localhost:9092", "broker list", "BV broker list defined", "", "Inspect runtime config and broker endpoints", "SparkKafkaStreaming Scala sources", "High", "", "Current code assumes a local broker."],
        ["ARCH-03", "Architecture", "Stream micro-batch interval", "Controls monitoring cadence and alert freshness.", 10, "s", "<= 10 s effective batch interval", "", "Observe StreamingContext and runtime logs", "KafkaConsumerRF.scala", "High", "", "Configured, but not yet experimentally benchmarked in the paper."],
        ["ARCH-04", "Architecture", "RF pipeline artifact", "Consumer loads a saved Spark ML pipeline instead of calling an external inference API.", "SparkKafkaStreaming/model_artifacts/random_forest_smote_standard_pipeline", "path", "Artifact exists and loads successfully during BV run", "", "Health-check model load on startup", "KafkaConsumerRF.scala; RandomForestModelExport.scala", "High", "", "Streaming inference now uses the saved Random Forest pipeline."],
        ["ARCH-05", "Architecture", "Model/data path portability", "Relative env-based paths should remain portable across local BV environments.", "Env-based relative paths", "state", "Externalized config before BV production sign-off", "", "Inspect startup config", "KafkaConsumerRF.scala; RandomForestModelExport.scala", "High", "", "Training data and model paths are now env-configurable relative paths."],
        ["ARCH-06", "Architecture", "Current distribution level", "Needed for honest BV reporting on 'how distributed' the system currently is.", "Single host, multi-core Spark local mode", "topology", "BV topology explicitly recorded", "", "Document node/broker/executor counts", "Derived from Spark config", "High", "", "Do not report this as multi-node distributed execution."],
        ["SIG-01", "Signal Coverage", "Predictive signal count", "Confirms the model sees the full expected sensor set.", len(feature_cols), "signals", f"{len(feature_cols)}/{len(feature_cols)} available in BV feed", "", "Compare incoming schema against Signal_Coverage sheet", "Data/*.csv header; report/main.tex", "High", "", "Training uses 18 numeric features; Timestamp excluded from model input."],
        ["SIG-02", "Signal Coverage", "Subsystem coverage", "Ensures all HRSS subsystems are represented.", len(subsystems), "subsystems", f"{len(subsystems)}/{len(subsystems)} available", "", "Check subsystem presence in incoming schema", "report/main.tex", "High", "", ", ".join(subsystems)],
        ["SIG-03", "Signal Coverage", "Signal family coverage", "Confirms current, power, and voltage channels are all present.", f"current={current_count}, power={power_count}, voltage={voltage_count}", "counts", "All three channel families present", "", "Schema check", "Data/*.csv header; report/main.tex", "High", "", "A missing family would invalidate model comparability."],
        ["SIG-04", "Signal Coverage", "Timestamp availability", "Needed for event sequencing and latency measurement.", "Yes", "boolean", "Present in BV feed", "", "Schema check", "Data/*.csv header", "High", "", "Timestamp is used for process interpretation and latency instrumentation."],
        ["SIG-05", "Signal Coverage", "Labels availability", "Needed for offline BV scoring against ground truth.", "Yes in study datasets", "boolean", "Present for validation datasets", "", "Schema/data check", "Data/*.csv header", "High", "", "Online production feeds may not include labels; offline BV feeds should."],
        ["DATA-01", "Data Scope", "Raw standard dataset size", "Reference baseline for study-side comparison.", 23645, "records", "Use as study baseline", "", "Compare against BV validation subset", "report/main.tex", "Medium", "", "17,975 normal + 5,670 anomalous."],
        ["DATA-02", "Data Scope", "Raw optimized dataset size", "Reference baseline for study-side comparison.", 19634, "records", "Use as study baseline", "", "Compare against BV validation subset", "report/main.tex", "Medium", "", "15,117 normal + 4,517 anomalous."],
        ["DATA-03", "Data Scope", "Raw standard anomaly rate", "Useful to compare operational class imbalance.", 0.2398, "ratio", "Comparable BV operating slice selected", "", "Compute anomaly_count / total_count", "report/main.tex", "Medium", "", "23.98%."],
        ["DATA-04", "Data Scope", "Raw optimized anomaly rate", "Useful to compare operational class imbalance.", 0.2301, "ratio", "Comparable BV operating slice selected", "", "Compute anomaly_count / total_count", "report/main.tex", "Medium", "", "23.01%."],
        ["DET-01", "Detection Quality", "Raw standard F1", "Core anomaly-capture KPI on the standard operating behaviour.", 0.9630, "F1", ">= 0.95", "", "Score predictions against ground truth", "report/main.tex", "High", "", "Best raw model = Random Forest."],
        ["DET-02", "Detection Quality", "Raw optimized F1", "Core anomaly-capture KPI on the optimized operating behaviour.", 0.9539, "F1", ">= 0.95", "", "Score predictions against ground truth", "report/main.tex", "High", "", "Best raw model = Random Forest."],
        ["DET-03", "Detection Quality", "Raw standard precision", "Shows how clean the flagged review set is.", 0.9635, "ratio", ">= 0.95", "", "TP / (TP + FP)", "report/main.tex", "High", "", "Helps quantify review efficiency."],
        ["DET-04", "Detection Quality", "Raw optimized precision", "Shows how clean the flagged review set is.", 0.9579, "ratio", ">= 0.95", "", "TP / (TP + FP)", "report/main.tex", "High", "", "Helps quantify review efficiency."],
        ["DET-05", "Detection Quality", "Raw standard recall", "Shows how many anomalous signals/events are captured.", 0.9626, "ratio", ">= 0.95", "", "TP / (TP + FN)", "report/main.tex", "High", "", "Primary answer to 'signals ne kadar yakalandı?'"],
        ["DET-06", "Detection Quality", "Raw optimized recall", "Shows how many anomalous signals/events are captured.", 0.9500, "ratio", ">= 0.95", "", "TP / (TP + FN)", "report/main.tex", "High", "", "Primary answer to 'signals ne kadar yakalandı?'"],
        ["DET-07", "Detection Quality", "Raw standard normal-state FPR", "Controls false alarm burden on normal traffic.", 0.0115, "ratio", "<= 0.015", "", "FP / normal_count", "report/main.tex", "High", "", "1.15%."],
        ["DET-08", "Detection Quality", "Raw optimized normal-state FPR", "Controls false alarm burden on normal traffic.", 0.0125, "ratio", "<= 0.015", "", "FP / normal_count", "report/main.tex", "High", "", "1.25%."],
        ["DET-09", "Operational Value", "Combined review reduction", "Shows how much manual review load is compressed.", 0.7656, "ratio", ">= 0.75", "", "1 - review_count / total_count", "report/main.tex", "High", "", "76.56%."],
        ["DET-10", "Operational Value", "Illustrative avoided inspections", "Translates model performance into operator workload reduction.", 33135, "records", "Informational", "", "Total_count - review_count", "report/main.tex", "Medium", "", "Derived metric, not field-measured KPI."],
        ["DET-11", "Operational Value", "Illustrative saved review effort", "Operational translation of compressed review set.", 92.0, "staff-hours", "Informational", "", "Assumes 10 seconds per reviewed record", "report/main.tex", "Medium", "", "Illustrative only."],
        ["DET-12", "Detection Quality", "Best balanced offline F1", "Upper benchmark ceiling under SMOTE-balanced standard data.", 0.9902, "F1", ">= 0.98 if balanced BV replay is executed", "", "Offline replay scoring", "report/main.tex", "Medium", "", "Use only if BV includes balanced offline replay."],
        ["ROUTE-01", "Routing Logic", "Anomaly threshold", "Defines which scored records are published as anomaly alerts.", "> 0.8", "rule", "Applied exactly as specified or revised explicitly", "", "Check code and runtime behaviour", "KafkaAnomalyDetection.scala", "High", "", "Used for anomaly topic routing."],
        ["ROUTE-02", "Routing Logic", "Normal threshold", "Defines which scored records are published as normal confirmations.", "<= 0.2", "rule", "Applied exactly as specified or revised explicitly", "", "Check code and runtime behaviour", "KafkaAnomalyDetection.scala", "High", "", "Used for normal_data topic routing."],
        ["ROUTE-03", "Routing Logic", "Uncertain band handling", "Records between 0.2 and 0.8 currently have no dedicated routing path.", "0.2 < p <= 0.8 not routed", "state", "Decision documented before sign-off", "", "Inspect topic outputs and dropped/held records", "KafkaAnomalyDetection.scala; predictions.csv", "High", "", "Current implementation gap."],
        ["ROUTE-04", "Routing Logic", "Current sample scored records", "Reference size of the available scoring sample.", predictions["total"], "records", "Informational", "", "Count rows in predictions.csv", "SparkKafkaStreaming/predictions.csv", "Medium", "", "Current sample is from predictions.csv."],
        ["ROUTE-05", "Routing Logic", "Current sample anomaly-routed volume", "Shows how much data would flow to the anomaly topic with the current threshold.", predictions["anomaly"], "records", "Informational", "", "Count predictions > 0.8", "SparkKafkaStreaming/predictions.csv", "Medium", "", f"{predictions['anomaly_ratio']:.2%} of scored sample."],
        ["ROUTE-06", "Routing Logic", "Current sample normal-routed volume", "Shows how much data would flow to the normal topic with the current threshold.", predictions["normal"], "records", "Informational", "", "Count predictions <= 0.2", "SparkKafkaStreaming/predictions.csv", "Medium", "", f"{predictions['normal_ratio']:.2%} of scored sample."],
        ["ROUTE-07", "Routing Logic", "Current sample uncertain volume", "Quantifies the routing blind zone under the current threshold policy.", predictions["uncertain"], "records", "Should be explicitly handled or accepted", "", "Count 0.2 < predictions <= 0.8", "SparkKafkaStreaming/predictions.csv", "High", "", f"{predictions['uncertain_ratio']:.2%} of scored sample."],
        ["ROUTE-08", "Kafka Contract", "Consumer input topic", "Defines the topic consumed for model serving.", "model-input", "topic", "BV end-to-end topic contract approved", "", "Inspect consumer subscription", "KafkaConsumerRF.scala", "High", "", "Input topic."],
        ["ROUTE-09", "Kafka Contract", "Producer input topic alignment", "Raw-data producer writes directly to the consumer input topic.", "model-input", "topic", "Producer and consumer use the same input topic", "", "Inspect producer config", "KafkaProducerRF.scala", "High", "", "No topic mismatch should remain in the live BV path."],
        ["ROUTE-10", "Kafka Contract", "Anomaly publication topic", "Defines where anomaly alerts are published.", "anomalies3", "topic", "Topic exists and is monitored in BV", "", "Inspect Kafka output", "KafkaAnomalyDetection.scala", "High", "", "Alert topic."],
        ["ROUTE-11", "Kafka Contract", "Normal publication topic", "Defines where normal records are published.", "normal_data", "topic", "Topic exists and is monitored in BV", "", "Inspect Kafka output", "KafkaAnomalyDetection.scala", "Medium", "", "Normal routing topic."],
        ["PERF-01", "Runtime Performance", "Detection latency", "Measures event-to-score delay.", "TBD in BV", "ms", "Set by BV team", "", "event_ts to prediction_ts", "Not measured in paper", "High", "", "Must be instrumented in BV."],
        ["PERF-02", "Runtime Performance", "Kafka publish latency", "Measures score-to-topic delay.", "TBD in BV", "ms", "Set by BV team", "", "prediction_ts to kafka_send_ts", "Not measured in paper", "High", "", "Must be instrumented in BV."],
        ["PERF-03", "Runtime Performance", "Throughput", "Shows how many records can be processed per second.", "TBD in BV", "records/s", "Set by BV team", "", "Processed_count / wall_clock_time", "Not measured in paper", "High", "", "Needed for capacity discussion."],
        ["PERF-04", "Runtime Performance", "Consumer lag", "Shows whether stream processing keeps up with incoming load.", "TBD in BV", "records or ms", "Near-zero under test load", "", "Kafka lag observation", "Not measured in paper", "High", "", "Needed for sustained stream tests."],
        ["PERF-05", "Runtime Performance", "Alert delivery success", "Ensures anomaly alerts are not lost in transport.", "TBD in BV", "ratio", "1.00", "", "Delivered_alerts / expected_alerts", "Not measured in paper", "High", "", "Should be measured during scripted alert replay."],
        ["PERF-06", "Runtime Performance", "Dropped/ignored uncertain records", "Quantifies data not sent to any Kafka topic under current threshold rules.", "TBD in BV", "ratio", "0 if explicit uncertain routing is added; otherwise record and accept", "", "Uncertain_count / total_scored", "Derived from current logic", "High", "", "Current code leaves this band unmanaged."],
    ]
    add_rows(ws, test_rows)
    style_sheet(
        ws,
        widths={
            "A": 12,
            "B": 18,
            "C": 26,
            "D": 34,
            "E": 26,
            "F": 12,
            "G": 28,
            "H": 18,
            "I": 24,
            "J": 28,
            "K": 10,
            "L": 12,
            "M": 30,
        },
        input_cols=(8, 12, 13),
    )

    # Offline baselines
    ws = wb.create_sheet("Offline_Baselines")
    ws.append([
        "Dataset_Variant",
        "Sampling_Regime",
        "Best_Model",
        "Accuracy",
        "Precision",
        "Recall",
        "F1",
        "Review_Reduction",
        "Normal_FPR",
        "Notes",
    ])
    baseline_rows = [
        ["Raw standard", "Raw", "Random Forest", 0.9823, 0.9635, 0.9626, 0.9630, 0.7604, 0.0115, "Best raw standard operating behaviour result."],
        ["Raw optimized", "Raw", "Random Forest", 0.9789, 0.9579, 0.9500, 0.9539, 0.7718, 0.0125, "Best raw optimized operating behaviour result."],
        ["Undersampled standard", "Undersampling", "Random Forest", 0.9471, 0.9239, 0.9744, 0.9485, "", "", "Balanced offline benchmark."],
        ["Undersampled optimized", "Undersampling", "Random Forest", 0.9444, 0.9230, 0.9699, 0.9458, "", "", "Balanced offline benchmark."],
        ["SMOTE-balanced standard", "SMOTE", "Random Forest", 0.9902, 0.9882, 0.9922, 0.9902, "", "", "Best overall benchmark result."],
        ["SMOTE-balanced optimized", "SMOTE", "Random Forest", 0.9884, 0.9867, 0.9901, 0.9884, "", "", "Second-best overall benchmark result."],
        ["Regime mean F1", "Raw", "All models", "", "", "", 0.4596, "", "", "Mean F1 across all evaluated models."],
        ["Regime mean F1", "Undersampling", "All models", "", "", "", 0.7237, "", "", "Mean F1 across all evaluated models."],
        ["Regime mean F1", "SMOTE", "All models", "", "", "", 0.7643, "", "", "Mean F1 across all evaluated models."],
    ]
    add_rows(ws, baseline_rows)
    style_sheet(
        ws,
        widths={"A": 24, "B": 18, "C": 18, "D": 12, "E": 12, "F": 12, "G": 12, "H": 16, "I": 12, "J": 34},
    )

    # Process benefits
    ws = wb.create_sheet("Process_Benefits")
    ws.append(["Indicator", "Dataset_1_Standard", "Dataset_2_Optimized", "Relative_Change", "Notes"])
    process_rows = [
        ["Observed window from Timestamp", 18.851, 14.246, -0.2443, "Shorter observation window in optimized behaviour."],
        ["Mean aggregate power", 47678.63, 44537.42, -0.0659, "Lower mean aggregate power in optimized behaviour."],
        ["Median aggregate power", 48103.00, 43554.00, -0.0946, "Lower median aggregate power in optimized behaviour."],
        ["Aggregate energy proxy", 872424.88, 631725.73, -0.2759, "Comparative proxy, not billing-grade estimate."],
        ["Normal-state mean aggregate power", 47702.76, 43786.11, -0.0821, "Lower nominal load in optimized behaviour."],
        ["Anomaly rate", 0.2398, 0.2301, -0.0406, "Operating-context difference."],
        ["Delta anom-normal aggregate power", -100.64, 3265.72, "n/a", "Anomaly contrast becomes clearer in optimized behaviour."],
    ]
    add_rows(ws, process_rows)
    style_sheet(ws, widths={"A": 28, "B": 18, "C": 18, "D": 16, "E": 40})

    # Signal coverage
    ws = wb.create_sheet("Signal_Coverage")
    ws.append(["Column", "Subsystem", "Signal_Type", "Used_In_Model", "Used_In_Process_Analysis", "Description"])
    for col in header:
        subsystem, signal_type, desc = signal_meta(col)
        used_in_model = "No" if col in {"Timestamp", "Labels"} else "Yes"
        used_in_process = "Yes"
        if col == "Labels":
            used_in_process = "Yes"
        ws.append([col, subsystem, signal_type, used_in_model, used_in_process, desc])
    style_sheet(ws, widths={"A": 24, "B": 12, "C": 14, "D": 14, "E": 22, "F": 40})

    # Config map
    ws = wb.create_sheet("Config_Map")
    ws.append(["Config_Area", "Parameter", "Current_Value", "Unit", "Source", "Why_It_Matters"])
    config_rows = [
        ["Execution", "Spark master", "local[*]", "mode", "KafkaAnomalyDetection.scala; KafkaConsumerRF.scala", "Defines current execution topology."],
        ["Execution", "Micro-batch interval", 10, "s", "KafkaConsumerRF.scala", "Controls stream cadence."],
        ["Kafka", "Bootstrap servers", "localhost:9092", "endpoint", "SparkKafkaStreaming Scala sources", "Current broker assumption."],
        ["Kafka", "Consumer group", "rf-group", "group id", "KafkaConsumerRF.scala", "Needed for lag and offset tracking."],
        ["Kafka", "Consumer input topic", "model-input", "topic", "KafkaConsumerRF.scala", "Model-serving input."],
        ["Kafka", "Raw-data producer topic", "model-input", "topic", "KafkaProducerRF.scala", "Producer publishes directly to the consumer input topic."],
        ["Kafka", "Anomaly output topic", "anomalies3", "topic", "KafkaAnomalyDetection.scala", "Alert channel."],
        ["Kafka", "Normal output topic", "normal_data", "topic", "KafkaAnomalyDetection.scala", "Normal-state channel."],
        ["Scoring", "Anomaly threshold", "> 0.8", "rule", "KafkaAnomalyDetection.scala", "Alert routing condition."],
        ["Scoring", "Normal threshold", "<= 0.2", "rule", "KafkaAnomalyDetection.scala", "Normal routing condition."],
        ["Scoring", "Uncertain band", "0.2 < p <= 0.8", "rule", "Derived from thresholds", "Currently no dedicated routing path."],
        ["Inference", "RF pipeline artifact", "SparkKafkaStreaming/model_artifacts/random_forest_smote_standard_pipeline", "path", "KafkaConsumerRF.scala; RandomForestModelExport.scala", "Required model-serving dependency."],
        ["Environment", "CPU execution", "Yes", "boolean", "report/main.tex", "Paper results were CPU-based."],
        ["Environment", "RAM", 42.6, "GB", "report/main.tex", "Reference benchmark workstation memory."],
        ["Environment", "Python version", "3.12.3", "version", "report/main.tex", "Reproducibility baseline."],
        ["Environment", "OpenJDK version", "17.0.18", "version", "report/main.tex", "Reproducibility baseline."],
        ["Environment", "Scala target", "2.11.8", "version", "report/main.tex; build.sbt", "Reproducibility baseline."],
        ["Environment", "Spark target", "2.3.1", "version", "report/main.tex; build.sbt", "Reproducibility baseline."],
    ]
    add_rows(ws, config_rows)
    style_sheet(ws, widths={"A": 16, "B": 24, "C": 22, "D": 12, "E": 28, "F": 42})

    # Metric definitions
    ws = wb.create_sheet("Metric_Definitions")
    ws.append(["Metric", "Definition", "Why_BV_Should_Check_It", "Suggested_Interpretation"])
    metric_rows = [
        ["Precision", "TP / (TP + FP)", "Shows how clean the flagged alarm set is.", "Higher is better; directly tied to review efficiency."],
        ["Recall", "TP / (TP + FN)", "Shows how many anomalous signals/events are captured.", "Primary KPI for 'sinyalleri ne kadar yakaladık'."],
        ["F1", "2 * Precision * Recall / (Precision + Recall)", "Balances missed anomalies and false alarms.", "Best single offline detection KPI here."],
        ["Normal-state FPR", "FP / normal_count", "Shows false alarm burden on normal traffic.", "Keep low to avoid alert fatigue."],
        ["Review reduction", "1 - review_count / total_count", "Shows how much manual review is reduced.", "Operational value KPI."],
        ["Detection latency", "prediction_ts - event_ts", "Shows how quickly the monitoring stack reacts.", "Must be instrumented in BV; not measured in the paper."],
        ["Kafka publish latency", "kafka_send_ts - prediction_ts", "Shows transport delay after scoring.", "Must be instrumented in BV."],
        ["Throughput", "processed_records / elapsed_time", "Shows capacity under load.", "Needed for production sizing."],
        ["Consumer lag", "Broker offset - consumer offset", "Shows whether stream processing keeps up.", "Near-zero under nominal load is preferred."],
        ["Signal coverage", "available_signals / expected_signals", "Shows whether the deployed feed matches the trained schema.", "Must be 18/18 for direct model comparability."],
        ["Distribution level", "Documented node/broker/executor layout", "Lets BV state honestly how distributed the tested stack really is.", "Current repo = single host, multi-core local mode."],
    ]
    add_rows(ws, metric_rows)
    style_sheet(ws, widths={"A": 20, "B": 34, "C": 34, "D": 34})

    # Visual emphasis for notes sheets
    for sheet_name in ["Config_Map", "Offline_Baselines", "Process_Benefits", "Signal_Coverage", "Metric_Definitions"]:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if cell.column <= 3:
                    continue
                cell.fill = NOTE_FILL

    ws = wb["Offline_Baselines"]
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=10).fill = NOTE_FILL

    ws = wb["BV_Test_Cases"]
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=5).fill = NOTE_FILL
        ws.cell(row=row, column=10).fill = NOTE_FILL
        ws.cell(row=row, column=11).fill = GOOD_FILL

    return wb


if __name__ == "__main__":
    wb = build_workbook()
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    print(f"Wrote {OUT_PATH}")
